import os
import tempfile
import zipfile
from collections import defaultdict

from qgis.core import (
    QgsField,
    QgsFeatureRequest,
    QgsLayerTreeGroup,
    QgsLayerTreeLayer,
    QgsMapLayerType,
    QgsProject,
    QgsVectorDataProvider,
    QgsVectorFileWriter,
    QgsVectorLayer,
    QgsWkbTypes,
)
from qgis.gui import QgsProjectionSelectionDialog
from qgis.PyQt.QtCore import QDate, QDateTime, QSize, QVariant, Qt
from qgis.PyQt.QtGui import QTextCursor
from qgis.PyQt.QtWidgets import (
    QAbstractItemView,
    QAbstractSpinBox,
    QApplication,
    QCheckBox,
    QComboBox,
    QDialog,
    QFileDialog,
    QFormLayout,
    QGridLayout,
    QFrame,
    QGroupBox,
    QHBoxLayout,
    QLabel,
    QLineEdit,
    QListWidget,
    QListWidgetItem,
    QMainWindow,
    QMessageBox,
    QInputDialog,
    QPushButton,
    QSpinBox,
    QSplitter,
    QStackedWidget,
    QStyle,
    QTableWidget,
    QTableWidgetItem,
    QTextEdit,
    QVBoxLayout,
    QWidget,
    QHeaderView,
)

# --- EXCEL: xuất/nhập thuộc tính (nhúng trong file, không cần thêm file .py khác; cần openpyxl) ---
try:
    from qgis.core import NULL as QGIS_NULL
except ImportError:
    QGIS_NULL = object()

try:
    from openpyxl import Workbook, load_workbook
except ImportError:
    Workbook = None
    load_workbook = None

_FID_HEADER = "qgis_fid"

_QUICK_LAYER_NAME_OPTIONS = [
    "BanKinhBoViaBanKinhTimDuong_P",
    "BoViaDaiPhanCach_L",
    "CaoDoCongTNM_P",
    "CaoDoCongThoatTNT_P",
    "CaoDoNen_P",
    "CayXanh_P",
    "ChiGioiDuongDo_L",
    "ChiGioiXayDung_L",
    "ChucNangCongTrinh_P",
    "ChucNangSuDungDat_A",
    "CongTrinhCBKT_A",
    "CongTrinhCBKT_L",
    "CongTrinhCBKT_P",
    "CongTrinhCapDien_A",
    "CongTrinhCapDien_P",
    "CongTrinhChieuSang_P",
    "CongTrinhGiaoThong_A",
    "CongTrinhGiaoThong_L",
    "CongTrinhGiaoThong_P",
    "CongTrinhNangLuong_A",
    "CongTrinhNangLuong_P",
    "CongTrinhNgam_A",
    "CongTrinhNgam_L",
    "CongTrinhNgam_P",
    "CongTrinhTNTvaVSMT_A",
    "CongTrinhTNTvaVSMT_L",
    "CongTrinhTNTvaVSMT_P",
    "CongTrinhThongTin_A",
    "CongTrinhThongTin_P",
    "CongTrinh_A",
    "CongTrinh_L",
    "CongtrinhCapNuocPCCC_A",
    "CongtrinhCapNuocPCCC_P",
    "DanhGiaMoiTruong_A",
    "DanhGiaMoiTruong_L",
    "DanhGiaMoiTruong_P",
    "DiemDauNoi_P",
    "DiemNhanChinh_P",
    "DiemQuanTrac_P",
    "DiemToaDoTimDuongChuyenHuongTimDuong_P",
    "DongMucThietKe_L",
    "DuAnLienQuan_A",
    "GiaiPhapBaoVeMoiTruong_A",
    "GiaiPhapBaoVeMoiTruong_L",
    "GiaiPhapBaoVeMoiTruong_P",
    "HanhLangAnToan_L",
    "HuongDi_L",
    "HuongThoatNuocMua_L",
    "HuongThoatNuocThai_L",
    "KhongGianKTCQ_A",
    "KhongGianKTCQ_L",
    "KhuVucPhoiCanh_A",
    "MangLuoiCapNuoc_L",
    "MangLuoiCapThongTin_L",
    "MangLuoiChieuSang_L",
    "MangLuoiGiaoThongDuongBo_A",
    "MangLuoiGiaoThongDuongBo_L",
    "MangLuoiGiaoThongDuongKhong_L",
    "MangLuoiGiaoThongDuongSat_L",
    "MangLuoiGiaoThongDuongThuy_L",
    "MangLuoiNangLuong_L",
    "MangLuoiPhanPhoiDien_L",
    "MangLuoiThoatNuocMua_L",
    "MangLuoiThoatNuocThai_L",
    "MangLuoiTuyenBus_L",
    "MatCatNgang_L",
    "MatNuoc_A",
    "MocGioiQuyHoach_A",
    "MocGioiQuyHoach_L",
    "MocGioiQuyHoach_P",
    "NutTinhToanTNT_P",
    "PhanKhuQuyHoach_A",
    "PhanLuuThoatNuocMua_L",
    "PhanLuuThoatNuocThai_L",
    "PhanOQuyHoach_A",
    "PhanVungCapDien_A",
    "PhanVungCapNuoc_A",
    "PhanVungDanhGia_A",
    "PhanVungLuuVuc_A",
    "PhanVungPhucVu_A",
    "PhanVungSDDkhac_A",
    "PhanVungSanNen_A",
    "RanhGioiHanhChinh_L",
    "RanhGioiQuyHoach_A",
    "TenDonViHanhChinh_P",
    "ThongTinSanNen_P",
    "TuyenTKDT_L",
]


def _openpyxl_available():
    return Workbook is not None and load_workbook is not None


def sanitize_filename(value):
    text = str(value).strip()
    if not text:
        return "layer"
    for ch in '<>:"/\\|?*':
        text = text.replace(ch, "_")
    text = text.replace(" ", "_").replace("-", "_")
    return text.strip("._") or "layer"


def _excel_scalar(val):
    if val is None or val == QGIS_NULL:
        return None
    if isinstance(val, QVariant) and val.isNull():
        return None
    if isinstance(val, QDateTime):
        if val.isValid():
            return val.toPyDateTime()
        return None
    if isinstance(val, QDate):
        if val.isValid():
            return val.toPyDate()
        return None
    return val


def _write_layer_workbook(layer):
    wb = Workbook()
    ws = wb.active
    ws.title = "thuoc_tinh"
    fields = layer.fields()
    headers = [_FID_HEADER] + [fields.at(i).name() for i in range(fields.count())]
    ws.append(headers)
    for feat in layer.getFeatures():
        row = [feat.id()]
        for i in range(fields.count()):
            row.append(_excel_scalar(feat.attribute(i)))
        ws.append(row)
    return wb


def _save_workbook(wb, path):
    os.makedirs(os.path.dirname(path) or ".", exist_ok=True)
    wb.save(path)


def export_layers_attributes_excel(layers, output_path):
    if not _openpyxl_available():
        return "", "Thiếu thư viện openpyxl. Cài trong Python của QGIS: pip install openpyxl"
    if not layers:
        return "", "Không có lớp nào được chọn."
    n = len(layers)
    if n == 1:
        layer = layers[0]
        path = output_path.strip().rstrip("/\\")
        if not path:
            path = os.path.join(".", sanitize_filename(layer.name()) + ".xlsx")
        elif os.path.isdir(path):
            path = os.path.join(path, sanitize_filename(layer.name()) + ".xlsx")
        else:
            base_dir, fname = os.path.split(path)
            stem, ext = os.path.splitext(fname)
            if not stem:
                path = os.path.join(base_dir or ".", sanitize_filename(layer.name()) + ".xlsx")
        if not path.lower().endswith(".xlsx"):
            path = path + ".xlsx" if not path.endswith((".xlsx", ".zip")) else path
        if not path.lower().endswith(".xlsx"):
            path = os.path.splitext(path)[0] + ".xlsx"
        wb = _write_layer_workbook(layer)
        _save_workbook(wb, path)
        return path, None
    zip_path = output_path.strip().rstrip("/\\")
    if not zip_path:
        zip_path = os.path.join(".", f"{sanitize_filename(layers[0].name())}_{n}lop.zip")
    elif os.path.isdir(zip_path):
        zip_path = os.path.join(zip_path, f"{sanitize_filename(layers[0].name())}_{n}lop.zip")
    if not zip_path.lower().endswith(".zip"):
        zip_path = os.path.splitext(zip_path)[0] + ".zip"
    used_names = {}
    with tempfile.TemporaryDirectory(prefix="hsg_excel_") as tmp:
        files_to_zip = []
        for layer in layers:
            base = sanitize_filename(layer.name())
            k = used_names.get(base, 0)
            used_names[base] = k + 1
            fname = f"{base}.xlsx" if k == 0 else f"{base}_{k + 1}.xlsx"
            fpath = os.path.join(tmp, fname)
            wb = _write_layer_workbook(layer)
            _save_workbook(wb, fpath)
            files_to_zip.append((fpath, fname))
        with zipfile.ZipFile(zip_path, "w", zipfile.ZIP_DEFLATED) as zf:
            for fpath, arcname in files_to_zip:
                zf.write(fpath, arcname)
    return zip_path, None


def _header_map(ws):
    row = next(ws.iter_rows(min_row=1, max_row=1, values_only=True), None)
    if not row:
        return {}
    return {str(c).strip(): i for i, c in enumerate(row) if c is not None and str(c).strip()}


def _apply_sheet_to_layer(layer, ws):
    hmap = _header_map(ws)
    if _FID_HEADER not in hmap:
        return -1
    fid_col = hmap[_FID_HEADER]
    fields = layer.fields()
    field_names = [fields.at(i).name() for i in range(fields.count())]
    col_to_field_idx = {}
    for name in field_names:
        for hdr, cidx in hmap.items():
            if hdr == _FID_HEADER:
                continue
            if hdr == name or hdr == name[:10]:
                fi = fields.indexOf(name)
                if fi >= 0:
                    col_to_field_idx[cidx] = fi
                break
    layer.startEditing()
    pr = layer.dataProvider()
    updated = 0
    rows = list(ws.iter_rows(min_row=2, values_only=True))
    for row in rows:
        if row is None or fid_col >= len(row):
            continue
        raw_fid = row[fid_col]
        if raw_fid is None or str(raw_fid).strip() == "":
            continue
        try:
            fid = int(raw_fid)
        except (TypeError, ValueError):
            continue
        if not layer.getFeature(fid).isValid():
            continue
        attrs = {}
        for cidx, fidx in col_to_field_idx.items():
            if cidx >= len(row):
                continue
            val = row[cidx]
            attrs[fidx] = val
        if attrs:
            pr.changeAttributeValues({fid: attrs})
            updated += 1
    if not layer.commitChanges():
        layer.rollBack()
        return -2
    return updated


def import_layers_attributes_excel(layers, input_path, log_print=print):
    if not _openpyxl_available():
        return "Thiếu thư viện openpyxl. Cài trong Python của QGIS: pip install openpyxl"
    if not layers:
        return "Không có lớp nào được chọn."
    if not input_path or not os.path.isfile(input_path):
        return "File không tồn tại."
    ext = os.path.splitext(input_path)[1].lower()
    if ext == ".zip":
        groups = defaultdict(list)
        for lyr in layers:
            groups[sanitize_filename(lyr.name())].append(lyr)
        with tempfile.TemporaryDirectory(prefix="hsg_excel_in_") as tmp:
            with zipfile.ZipFile(input_path, "r") as zf:
                zf.extractall(tmp)
            stem_to_path = {}
            for root, _, files in os.walk(tmp):
                for fn in files:
                    if not fn.lower().endswith(".xlsx"):
                        continue
                    stem = os.path.splitext(fn)[0]
                    stem_to_path[stem] = os.path.join(root, fn)
            matched = 0
            for base, lyr_list in groups.items():
                for i, layer in enumerate(lyr_list):
                    stem_key = base if i == 0 else f"{base}_{i + 1}"
                    fpath = stem_to_path.get(stem_key)
                    if not fpath:
                        log_print(
                            f"  − Trong ZIP không có file tương ứng lớp «{layer.name()}» (mong đợi: {stem_key}.xlsx)."
                        )
                        continue
                    wb = load_workbook(fpath, read_only=True, data_only=True)
                    try:
                        ws = wb.active
                        n = _apply_sheet_to_layer(layer, ws)
                        if n == -2:
                            log_print(
                                f"  − Không lưu được thay đổi cho lớp {layer.name()} (từ {os.path.basename(fpath)})."
                            )
                        elif n < 0:
                            log_print(f"  − File {os.path.basename(fpath)} thiếu cột bắt buộc '{_FID_HEADER}'.")
                        else:
                            matched += 1
                            log_print(
                                f"  + Đã cập nhật {n} đối tượng từ {os.path.basename(fpath)} → {layer.name()}"
                            )
                    finally:
                        wb.close()
            if matched == 0:
                return "Không khớp được file .xlsx nào trong ZIP với các lớp đã chọn (tên file = tên lớp đã làm sạch)."
        return None
    if ext != ".xlsx":
        return "Chỉ hỗ trợ file .xlsx hoặc .zip."
    wb = load_workbook(input_path, read_only=True, data_only=True)
    try:
        ws = wb.active
        if len(layers) == 1:
            n = _apply_sheet_to_layer(layers[0], ws)
            if n == -2:
                return f"Không lưu được thay đổi cho lớp {layers[0].name()}."
            if n < 0:
                return f"File thiếu cột bắt buộc '{_FID_HEADER}'."
            log_print(f"  + Đã cập nhật {n} đối tượng cho lớp {layers[0].name()}.")
            return None
        stem = sanitize_filename(os.path.splitext(os.path.basename(input_path))[0])
        target = None
        for lyr in layers:
            if sanitize_filename(lyr.name()) == stem or lyr.name() == stem:
                target = lyr
                break
        if target is None:
            return (
                "Nhiều lớp được chọn: hãy dùng file ZIP từ bước xuất, hoặc chọn đúng một lớp, "
                "hoặc đặt tên file .xlsx trùng tên lớp (đã làm sạch)."
            )
        n = _apply_sheet_to_layer(target, ws)
        if n == -2:
            return f"Không lưu được thay đổi cho lớp {target.name()}."
        if n < 0:
            return f"File thiếu cột bắt buộc '{_FID_HEADER}'."
        log_print(f"  + Đã cập nhật {n} đối tượng cho lớp {target.name()}.")
        return None
    finally:
        wb.close()


# --- FIX TRUNCATION TEN LAYER DXF (OGR "Layer" HAY BI CAT 10 KY TU) ---
def _extract_dxf_layer_names(file_path: str):
    """
    Trả về list tên layer đầy đủ trong bảng LAYER của DXF (ASCII).
    """
    try:
        with open(file_path, "rb") as f:
            head = f.read(32)
        if b"AutoCAD Binary DXF" in head:
            return {}

        with open(file_path, "r", encoding="utf-8", errors="ignore") as f:
            lines = [ln.rstrip("\r\n") for ln in f]

        in_layer_table = False
        pending_layer_entity = False
        full_names = []
        i = 0
        n = len(lines)
        while i + 1 < n:
            code = lines[i].strip()
            value = lines[i + 1].strip()

            if code == "0" and value == "TABLE":
                if i + 3 < n and lines[i + 2].strip() == "2" and lines[i + 3].strip().upper() == "LAYER":
                    in_layer_table = True
                    i += 4
                    continue

            if in_layer_table and code == "0" and value == "ENDTAB":
                in_layer_table = False
                pending_layer_entity = False
                i += 2
                continue

            if in_layer_table and code == "0" and value == "LAYER":
                pending_layer_entity = True
                i += 2
                continue

            if in_layer_table and pending_layer_entity and code == "2":
                if value:
                    full_names.append(value)
                pending_layer_entity = False
                i += 2
                continue

            i += 2

        return full_names
    except Exception:
        return []


def _restore_layer_name(cad_layer_str: str, dxf_full_layer_names):
    if not dxf_full_layer_names:
        return cad_layer_str

    if cad_layer_str in dxf_full_layer_names:
        return cad_layer_str

    matches = [n for n in dxf_full_layer_names if n.startswith(cad_layer_str)]
    if len(matches) == 1:
        return matches[0]

    key10 = cad_layer_str[:10]
    matches10 = [n for n in dxf_full_layer_names if n.startswith(key10)]
    if len(matches10) == 1:
        return matches10[0]

    return cad_layer_str

# --- TUONG THICH ENUM GIUA CAC PHIEN BAN PYQT/QGIS ---
try:
    ORIENTATION_VERTICAL = Qt.Orientation.Vertical
except AttributeError:
    ORIENTATION_VERTICAL = Qt.Vertical
try:
    ORIENTATION_HORIZONTAL = Qt.Orientation.Horizontal
except AttributeError:
    ORIENTATION_HORIZONTAL = Qt.Horizontal

try:
    CHK_CHECKED = Qt.CheckState.Checked
    CHK_UNCHECKED = Qt.CheckState.Unchecked
except AttributeError:
    CHK_CHECKED = Qt.Checked
    CHK_UNCHECKED = Qt.Unchecked

try:
    FLAG_CHECKABLE = Qt.ItemFlag.ItemIsUserCheckable
except AttributeError:
    FLAG_CHECKABLE = Qt.ItemIsUserCheckable

try:
    USER_ROLE = Qt.ItemDataRole.UserRole
except AttributeError:
    USER_ROLE = Qt.UserRole

try:
    TEXT_CURSOR_END = QTextCursor.MoveOperation.End
except AttributeError:
    TEXT_CURSOR_END = QTextCursor.End

try:
    ALIGN_VCENTER = Qt.AlignmentFlag.AlignVCenter
except AttributeError:
    ALIGN_VCENTER = Qt.AlignVCenter

try:
    TABLE_NO_EDIT_TRIGGERS = QAbstractItemView.EditTrigger.NoEditTriggers
except AttributeError:
    TABLE_NO_EDIT_TRIGGERS = QAbstractItemView.NoEditTriggers

try:
    TABLE_SELECT_ROWS = QAbstractItemView.SelectionBehavior.SelectRows
except AttributeError:
    TABLE_SELECT_ROWS = QAbstractItemView.SelectRows

try:
    TABLE_EXT_SELECTION = QAbstractItemView.SelectionMode.ExtendedSelection
except AttributeError:
    TABLE_EXT_SELECTION = QAbstractItemView.ExtendedSelection

try:
    SPIN_NO_BUTTONS = QAbstractSpinBox.ButtonSymbols.NoButtons
except AttributeError:
    SPIN_NO_BUTTONS = QAbstractSpinBox.NoButtons

try:
    HEADER_STRETCH = QHeaderView.ResizeMode.Stretch
except AttributeError:
    HEADER_STRETCH = QHeaderView.Stretch

try:
    SCROLLBAR_ALWAYS_OFF = Qt.ScrollBarPolicy.ScrollBarAlwaysOff
except AttributeError:
    SCROLLBAR_ALWAYS_OFF = Qt.ScrollBarAlwaysOff
try:
    SCROLLBAR_AS_NEEDED = Qt.ScrollBarPolicy.ScrollBarAsNeeded
except AttributeError:
    SCROLLBAR_AS_NEEDED = Qt.ScrollBarAsNeeded

# Thương hiệu / bản quyền (hiển thị trên giao diện)
_DEV_BRAND = "LEDAT"
_COPYRIGHT_YEAR = "2026"


class HoSoGISWindow(QMainWindow):
    def __init__(self, parent=None):
        super().__init__(parent)
        self.setWindowTitle("HoSoGIS")
        self.resize(1140, 860)
        self.setStyleSheet(self._build_style())
        self._build_ui()
        self.refresh_vector_layers()

    def _build_style(self):
        return """
        QMainWindow { background-color: #f1f5f9; }
        QWidget {
            color: #334155;
            font-size: 12px;
        }
        QFrame#headerCard {
            background: #ffffff;
            border: 1px solid #e2e8f0;
            border-radius: 8px;
        }
        QLabel#title {
            font-size: 26px;
            font-weight: 800;
            color: #0f172a;
            letter-spacing: 0.02em;
        }
        QLabel#titleKicker {
            color: #2563eb;
            font-size: 10px;
            font-weight: 700;
            letter-spacing: 0.12em;
            text-transform: uppercase;
        }
        QLabel#subtitle { color: #64748b; font-size: 12px; }
        QFrame#footerBar {
            background: #ffffff;
            border: 1px solid #e2e8f0;
            border-radius: 8px;
            min-height: 40px;
        }
        QLabel#footerBrand {
            color: #2563eb;
            font-size: 12px;
            font-weight: 700;
            letter-spacing: 0.08em;
        }
        QLabel#footerCopyright {
            color: #64748b;
            font-size: 11px;
            font-weight: 500;
        }
        QLabel#footerDot {
            color: #94a3b8;
            font-size: 11px;
            padding: 0 6px;
        }
        QGroupBox {
            background: #ffffff;
            border: 1px solid #e2e8f0;
            border-radius: 8px;
            margin-top: 10px;
            font-weight: 600;
            color: #0f172a;
        }
        QGroupBox::title {
            subcontrol-origin: margin;
            left: 10px;
            padding: 0 4px;
            color: #0f172a;
        }
        QLineEdit, QListWidget, QTextEdit {
            border: 1px solid #cbd5e1;
            border-radius: 8px;
            padding: 6px 8px;
            background: #ffffff;
            selection-background-color: #bfdbfe;
            selection-color: #0f172a;
        }
        QListWidget::item {
            padding: 4px 6px;
            border-radius: 4px;
            margin: 1px 0;
        }
        QListWidget::item:selected {
            background: #dbeafe;
            color: #0f172a;
        }
        QLineEdit:focus, QListWidget:focus, QTextEdit:focus {
            border: 1px solid #2563eb;
            background: #ffffff;
        }
        QTableWidget {
            border: 1px solid #cbd5e1;
            border-radius: 8px;
            background: #ffffff;
            gridline-color: #e2e8f0;
            selection-background-color: #dbeafe;
            selection-color: #0f172a;
        }
        QHeaderView::section {
            background: #eff6ff;
            color: #1e3a8a;
            border: none;
            border-right: 1px solid #dbeafe;
            border-bottom: 1px solid #dbeafe;
            padding: 6px 8px;
            font-weight: 600;
        }
        QLineEdit::placeholder {
            color: #94a3b8;
        }
        QComboBox, QSpinBox {
            border: 1px solid #cbd5e1;
            border-radius: 8px;
            padding: 5px 8px;
            background: #ffffff;
            min-height: 1.2em;
        }
        QComboBox:focus, QSpinBox:focus {
            border: 1px solid #2563eb;
        }
        QComboBox::drop-down {
            border: none;
            width: 22px;
        }
        QDialog#hosogisSubDialog {
            background-color: #f8fafc;
        }
        QDialog#hosogisSubDialog QLabel {
            color: #475569;
        }
        QTableWidget::item {
            background: #ffffff;
            padding: 2px 4px;
        }
        QTableWidget::item:selected {
            background: #dbeafe;
            color: #0f172a;
        }
        QPushButton {
            background-color: #2563eb;
            color: white;
            border: none;
            border-radius: 8px;
            padding: 7px 12px;
            font-weight: 600;
        }
        QPushButton:hover { background-color: #1d4ed8; }
        QPushButton:pressed { background-color: #1e40af; }
        QPushButton:disabled {
            background-color: #cbd5e1;
            color: #f8fafc;
        }
        QPushButton#ghost {
            background-color: #ffffff;
            color: #334155;
            border: 1px solid #cbd5e1;
        }
        QPushButton#ghost:hover {
            background-color: #f8fafc;
            border: 1px solid #94a3b8;
        }
        QFrame#moduleNavFrame {
            background: #e8edf4;
            border: 1px solid #d8dee9;
            border-radius: 10px;
            min-width: 196px;
            max-width: 220px;
        }
        QLabel#moduleNavTitle {
            color: #64748b;
            font-size: 10px;
            font-weight: 700;
            letter-spacing: 0.14em;
            text-transform: uppercase;
            padding: 2px 4px 4px 4px;
        }
        QListWidget#moduleNav {
            border: none;
            background: transparent;
            outline: none;
            padding: 4px 0;
        }
        QListWidget#moduleNav::item {
            min-height: 44px;
            padding: 10px 12px;
            margin: 4px 8px;
            border-radius: 8px;
            border: 1px solid transparent;
            color: #475569;
            font-weight: 600;
        }
        QListWidget#moduleNav::item:hover {
            background: #f1f5f9;
            border: 1px solid #cbd5e1;
            color: #1e293b;
        }
        QListWidget#moduleNav::item:selected {
            background: #ffffff;
            border: 1px solid #93c5fd;
            color: #1e40af;
        }
        QStackedWidget#moduleStack {
            background: #ffffff;
            border: 1px solid #e2e8f0;
            border-radius: 10px;
        }
        QLabel#sectionHint {
            color: #64748b;
            background: #f8fafc;
            border: 1px solid #e2e8f0;
            border-radius: 8px;
            padding: 6px 8px;
        }
        QCheckBox {
            spacing: 7px;
            color: #475569;
        }
        QCheckBox::indicator {
            width: 16px;
            height: 16px;
        }
        QSplitter#hosogisMainSplit::handle,
        QSplitter#hosogisMainSplit::handle:horizontal,
        QSplitter#hosogisMainSplit::handle:vertical {
            background: #f1f5f9;
            border: none;
            margin: 0px;
            padding: 0px;
        }
        QSplitter#hosogisMainSplit::handle:hover,
        QSplitter#hosogisMainSplit::handle:horizontal:hover,
        QSplitter#hosogisMainSplit::handle:vertical:hover {
            background: #eef1f6;
        }
        QSplitter#hosogisWorkSplit::handle {
            background: #e2e8f0;
        }
        QSplitter#hosogisWorkSplit::handle:hover {
            background: #cbd5e1;
        }
        QTextEdit#logView {
            background: #ffffff;
        }
        QMessageBox {
            background-color: #ffffff;
        }
        QMessageBox QLabel {
            color: #334155;
            font-size: 12px;
        }
        QMessageBox QPushButton {
            min-width: 88px;
            padding: 7px 12px;
            border-radius: 8px;
        }
        """

    def _build_ui(self):
        container = QWidget()
        self.setCentralWidget(container)

        root_layout = QVBoxLayout(container)
        root_layout.setContentsMargins(12, 12, 12, 12)
        root_layout.setSpacing(8)

        header = QFrame()
        header.setObjectName("headerCard")
        header_layout = QVBoxLayout(header)
        header_layout.setContentsMargins(14, 12, 14, 12)
        header_layout.setSpacing(1)
        lbl_kicker = QLabel("PLANNING DATA WORKSPACE")
        lbl_kicker.setObjectName("titleKicker")
        lbl_title = QLabel('HoSo<span style="color:#2563eb;">GIS</span>')
        lbl_title.setObjectName("title")
        lbl_sub = QLabel("Quản lý dữ liệu quy hoạch: nhập CAD, cập nhật thuộc tính, xuất GDB/GPKG")
        lbl_sub.setObjectName("subtitle")
        header_layout.addWidget(lbl_kicker)
        header_layout.addWidget(lbl_title)
        header_layout.addWidget(lbl_sub)
        root_layout.addWidget(header)

        body_split = QSplitter(ORIENTATION_VERTICAL)
        body_split.setObjectName("hosogisMainSplit")
        body_split.setChildrenCollapsible(False)
        body_split.setHandleWidth(4)
        body_split.addWidget(self._build_left_panel())
        body_split.addWidget(self._build_log_panel())
        body_split.setStretchFactor(0, 1)
        body_split.setStretchFactor(1, 0)
        body_split.setSizes([560, 200])
        root_layout.addWidget(body_split, 1)

        footer = QFrame()
        footer.setObjectName("footerBar")
        footer_layout = QHBoxLayout(footer)
        footer_layout.setContentsMargins(16, 10, 16, 10)
        footer_layout.setSpacing(0)
        lbl_footer_brand = QLabel(_DEV_BRAND)
        lbl_footer_brand.setObjectName("footerBrand")
        lbl_dot = QLabel("·")
        lbl_dot.setObjectName("footerDot")
        lbl_footer_copy = QLabel(
            f"Bản quyền phần mềm HoSoGIS © {_COPYRIGHT_YEAR} {_DEV_BRAND}. "
            "Không sao chép, phân phối hoặc sửa đổi mà không được phép."
        )
        lbl_footer_copy.setObjectName("footerCopyright")
        lbl_footer_copy.setWordWrap(True)
        footer_layout.addWidget(lbl_footer_brand, 0, ALIGN_VCENTER)
        footer_layout.addWidget(lbl_dot, 0, ALIGN_VCENTER)
        footer_layout.addWidget(lbl_footer_copy, 1, ALIGN_VCENTER)
        root_layout.addWidget(footer)

    def _nav_std_icon(self, std_py6, std_py5):
        sty = self.style()
        try:
            return sty.standardIcon(getattr(QStyle.StandardPixmap, std_py6))
        except AttributeError:
            return sty.standardIcon(getattr(QStyle, std_py5))

    def _build_left_panel(self):
        panel = QWidget()
        layout = QHBoxLayout(panel)
        layout.setContentsMargins(0, 0, 0, 0)
        layout.setSpacing(12)

        nav_wrap = QFrame()
        nav_wrap.setObjectName("moduleNavFrame")
        nav_wrap.setFixedWidth(208)
        nav_outer = QVBoxLayout(nav_wrap)
        nav_outer.setContentsMargins(10, 12, 10, 12)
        nav_outer.setSpacing(6)

        nav_heading = QLabel("Chức năng")
        nav_heading.setObjectName("moduleNavTitle")
        nav_outer.addWidget(nav_heading)

        self.module_nav = QListWidget()
        self.module_nav.setObjectName("moduleNav")
        self.module_nav.setIconSize(QSize(22, 22))
        self.module_nav.setSpacing(2)
        self.module_nav.setHorizontalScrollBarPolicy(SCROLLBAR_ALWAYS_OFF)
        self.module_nav.setVerticalScrollBarPolicy(SCROLLBAR_AS_NEEDED)
        self.module_nav.setUniformItemSizes(True)

        nav_entries = [
            (
                self._nav_std_icon("SP_DialogOpenButton", "SP_DialogOpenButton"),
                "Nhập CAD",
                "Nhập bản vẽ CAD (DXF/DWG) và tách lớp",
            ),
            (
                self._nav_std_icon("SP_FileDialogInfoView", "SP_FileDialogInfoView"),
                "Thuộc tính",
                "Cập nhật trường chuẩn quy hoạch cho layer",
            ),
            (
                self._nav_std_icon("SP_BrowserReload", "SP_BrowserReload"),
                "Đổi tên nhanh",
                "Đổi tên layer theo danh mục chuẩn",
            ),
            (
                self._nav_std_icon("SP_DialogSaveButton", "SP_DialogSaveButton"),
                "Nhập / Xuất",
                "Nhập GDB và xuất GDB / GPKG",
            ),
        ]
        for icon, title, tip in nav_entries:
            item = QListWidgetItem(icon, title)
            item.setToolTip(tip)
            self.module_nav.addItem(item)

        self.module_stack = QStackedWidget()
        self.module_stack.setObjectName("moduleStack")
        self.module_stack.addWidget(self._build_tab_import_cad())
        self.module_stack.addWidget(self._build_tab_attributes())
        self.module_stack.addWidget(self._build_tab_quick_rename())
        self.module_stack.addWidget(self._build_tab_export())

        self.module_nav.currentRowChanged.connect(self.module_stack.setCurrentIndex)
        self.module_nav.setCurrentRow(0)

        nav_outer.addWidget(self.module_nav, 1)
        layout.addWidget(nav_wrap, 0)
        layout.addWidget(self.module_stack, 1)
        return panel

    def _build_log_panel(self):
        panel = QWidget()
        layout = QVBoxLayout(panel)
        layout.setContentsMargins(0, 4, 0, 0)
        layout.setSpacing(6)

        top = QHBoxLayout()
        top.setSpacing(8)
        lbl = QLabel("Nhật ký xử lý")
        lbl.setStyleSheet("font-weight: 700; color: #0f172a; font-size: 13px;")
        hint = QLabel("Thông báo từ các thao tác hiển thị tại đây; có thể kéo thanh phía trên để chỉnh chiều cao.")
        hint.setObjectName("sectionHint")
        hint.setWordWrap(True)
        top.addWidget(lbl)
        top.addWidget(hint, 1)
        btn_clear = QPushButton("Xóa nhật ký")
        btn_clear.setObjectName("ghost")
        btn_clear.clicked.connect(self._clear_log)
        top.addWidget(btn_clear)
        layout.addLayout(top)

        self.log_edit = QTextEdit()
        self.log_edit.setObjectName("logView")
        self.log_edit.setReadOnly(True)
        self.log_edit.setMinimumHeight(110)
        layout.addWidget(self.log_edit, 1)
        return panel

    def _clear_log(self):
        self.log_edit.clear()

    def _build_tab_import_cad(self):
        tab = QWidget()
        tab_layout = QVBoxLayout(tab)
        tab_layout.setContentsMargins(10, 10, 10, 10)
        tab_layout.setSpacing(10)

        hint = QLabel("Nhập file CAD (ưu tiên DXF) và tách lớp theo hình học.")
        hint.setObjectName("sectionHint")
        hint.setWordWrap(True)
        tab_layout.addWidget(hint)

        box = QGroupBox("Nguồn dữ liệu CAD")
        box_layout = QVBoxLayout(box)
        row = QHBoxLayout()
        self.input_cad_path = QLineEdit()
        self.input_cad_path.setPlaceholderText("Chọn file CAD (khuyến nghị DXF)")
        btn_browse = QPushButton("Chọn file")
        btn_browse.setObjectName("ghost")
        btn_browse.clicked.connect(self.choose_cad_file)
        row.addWidget(self.input_cad_path, 1)
        row.addWidget(btn_browse)
        box_layout.addLayout(row)

        btn_run = QPushButton("Chạy nhập bản vẽ")
        btn_run.clicked.connect(self.import_and_split_cad)
        box_layout.addWidget(btn_run)
        tab_layout.addWidget(box)
        tab_layout.addStretch(1)
        return tab

    def _build_tab_attributes(self):
        tab = QWidget()
        tab_layout = QVBoxLayout(tab)
        tab_layout.setContentsMargins(10, 10, 10, 10)
        tab_layout.setSpacing(10)

        hint = QLabel(
            "1) Nhập giá trị chuẩn bên trái.\n"
            "2) Chọn layer bên phải — xem bảng trường phía dưới (có thể thêm / đổi tên / xóa cột).\n"
            "3) Bấm «Chạy cập nhật thuộc tính»."
        )
        hint.setObjectName("sectionHint")
        hint.setWordWrap(True)
        tab_layout.addWidget(hint)

        main_splitter = QSplitter(ORIENTATION_HORIZONTAL)
        main_splitter.setObjectName("hosogisWorkSplit")
        main_splitter.setChildrenCollapsible(False)
        main_splitter.setHandleWidth(5)

        form_box = QGroupBox("Thông tin thuộc tính")
        form_layout = QVBoxLayout(form_box)
        form_layout.setContentsMargins(12, 14, 12, 12)
        form_layout.setSpacing(10)
        form_grid = QGridLayout()
        form_grid.setHorizontalSpacing(12)
        form_grid.setVerticalSpacing(8)
        form_grid.setColumnStretch(1, 1)
        form_grid.setColumnStretch(3, 1)

        self.input_ma_tt = QLineEdit()
        self.input_ma_hs = QLineEdit()
        self.input_ma_dt = QLineEdit()
        self.input_ten_dt = QLineEdit()
        self.input_phan_loai = QLineEdit()
        self.input_ghi_chu = QLineEdit()

        def _lbl(text):
            w = QLabel(text)
            w.setAlignment(ALIGN_VCENTER)
            return w

        form_grid.addWidget(_lbl("Mã thông tin QH"), 0, 0)
        form_grid.addWidget(self.input_ma_tt, 0, 1)
        form_grid.addWidget(_lbl("Mã hồ sơ QH"), 0, 2)
        form_grid.addWidget(self.input_ma_hs, 0, 3)
        form_grid.addWidget(_lbl("Mã đối tượng"), 1, 0)
        form_grid.addWidget(self.input_ma_dt, 1, 1)
        form_grid.addWidget(_lbl("Tên đối tượng"), 1, 2)
        form_grid.addWidget(self.input_ten_dt, 1, 3)
        form_grid.addWidget(_lbl("Phân loại"), 2, 0)
        form_grid.addWidget(self.input_phan_loai, 2, 1, 1, 3)
        form_grid.addWidget(_lbl("Ghi chú"), 3, 0)
        form_grid.addWidget(self.input_ghi_chu, 3, 1, 1, 3)
        form_layout.addLayout(form_grid)

        self.chk_delete_old = QCheckBox("Xóa thuộc tính cũ trước khi cập nhật")
        form_layout.addWidget(self.chk_delete_old)
        form_layout.addStretch(1)
        main_splitter.addWidget(form_box)

        right_panel = QWidget()
        right_layout = QVBoxLayout(right_panel)
        right_layout.setContentsMargins(0, 0, 0, 0)
        right_layout.setSpacing(8)

        layer_box = QGroupBox("Layer áp dụng")
        layer_layout = QVBoxLayout(layer_box)
        layer_layout.setContentsMargins(12, 12, 12, 14)
        layer_layout.setSpacing(8)
        filter_row = QHBoxLayout()
        filter_row.setSpacing(8)
        self.input_layer_filter = QLineEdit()
        self.input_layer_filter.setPlaceholderText("Lọc nhanh theo tên layer...")
        self.input_layer_filter.textChanged.connect(self.filter_attribute_layers)
        self.lbl_layer_count = QLabel("Đã chọn: 0 | Hiển thị: 0")
        self.lbl_layer_count.setStyleSheet("color:#64748b; font-weight:600;")
        filter_row.addWidget(self.input_layer_filter, 1)
        filter_row.addWidget(self.lbl_layer_count)
        layer_layout.addLayout(filter_row)

        self.list_layers = QListWidget()
        self.list_layers.currentItemChanged.connect(self.on_attribute_layer_changed)
        self.list_layers.itemChanged.connect(self.update_selected_layer_count)
        self.list_layers.setAlternatingRowColors(False)
        self.list_layers.setMinimumHeight(96)
        layer_layout.addWidget(self.list_layers)

        btn_row = QHBoxLayout()
        btn_row.setSpacing(8)
        btn_select_all = QPushButton("Chọn tất cả")
        btn_select_all.setObjectName("ghost")
        btn_unselect_all = QPushButton("Bỏ chọn tất cả")
        btn_unselect_all.setObjectName("ghost")
        btn_refresh = QPushButton("Làm mới")
        btn_refresh.setObjectName("ghost")
        btn_select_all.clicked.connect(self.select_all_layers)
        btn_unselect_all.clicked.connect(self.unselect_all_layers)
        btn_refresh.clicked.connect(self.refresh_vector_layers)
        btn_row.addWidget(btn_select_all)
        btn_row.addWidget(btn_unselect_all)
        btn_row.addWidget(btn_refresh)
        layer_layout.addLayout(btn_row)

        excel_row = QHBoxLayout()
        excel_row.setSpacing(8)
        btn_excel_export = QPushButton("Xuất Excel…")
        btn_excel_import = QPushButton("Nhập Excel…")
        btn_excel_export.setObjectName("ghost")
        btn_excel_import.setObjectName("ghost")
        btn_excel_export.clicked.connect(self.export_attributes_excel)
        btn_excel_import.clicked.connect(self.import_attributes_excel)
        excel_row.addWidget(btn_excel_export)
        excel_row.addWidget(btn_excel_import)
        layer_layout.addLayout(excel_row)

        preview_box = QGroupBox("Xem trước thuộc tính layer đang chọn")
        preview_layout = QVBoxLayout(preview_box)
        preview_layout.setContentsMargins(12, 12, 12, 14)
        preview_layout.setSpacing(8)
        self.attr_preview_meta = QLabel("Chọn 1 layer để xem cấu trúc trường dữ liệu.")
        self.attr_preview_meta.setObjectName("sectionHint")
        self.attr_preview_meta.setWordWrap(True)
        preview_layout.addWidget(self.attr_preview_meta)
        self.attr_preview_table = QTableWidget(0, 3)
        self.attr_preview_table.setHorizontalHeaderLabels(
            ["Tên thuộc tính", "Kiểu dữ liệu", "Độ dài dữ liệu"]
        )
        self.attr_preview_table.setEditTriggers(TABLE_NO_EDIT_TRIGGERS)
        self.attr_preview_table.setSelectionBehavior(TABLE_SELECT_ROWS)
        self.attr_preview_table.setSelectionMode(TABLE_EXT_SELECTION)
        self.attr_preview_table.verticalHeader().setVisible(False)
        self.attr_preview_table.setAlternatingRowColors(False)
        self.attr_preview_table.setShowGrid(True)
        self.attr_preview_table.setMinimumHeight(120)
        preview_layout.addWidget(self.attr_preview_table)

        field_btn_row = QHBoxLayout()
        field_btn_row.setSpacing(8)
        btn_add_field = QPushButton("Thêm thuộc tính…")
        btn_rename_field = QPushButton("Đổi tên…")
        btn_delete_field = QPushButton("Xóa thuộc tính")
        btn_rename_field.setObjectName("ghost")
        btn_delete_field.setObjectName("ghost")
        btn_add_field.clicked.connect(self.add_attribute_field_manual)
        btn_rename_field.clicked.connect(self.rename_attribute_field_manual)
        btn_delete_field.clicked.connect(self.delete_attribute_fields_manual)
        field_btn_row.addWidget(btn_add_field)
        field_btn_row.addWidget(btn_rename_field)
        field_btn_row.addWidget(btn_delete_field)
        field_btn_row.addStretch(1)
        preview_layout.addLayout(field_btn_row)

        right_splitter = QSplitter(ORIENTATION_VERTICAL)
        right_splitter.setObjectName("hosogisWorkSplit")
        right_splitter.setChildrenCollapsible(False)
        right_splitter.setHandleWidth(5)
        right_splitter.addWidget(layer_box)
        right_splitter.addWidget(preview_box)
        right_splitter.setChildrenCollapsible(False)
        right_splitter.setSizes([300, 220])
        right_layout.addWidget(right_splitter, 1)

        main_splitter.addWidget(right_panel)
        main_splitter.setSizes([420, 700])
        tab_layout.addWidget(main_splitter, 1)

        btn_run = QPushButton("Chạy cập nhật thuộc tính")
        btn_run.clicked.connect(self.add_fields_and_data)
        tab_layout.addWidget(btn_run)
        return tab

    def _build_tab_quick_rename(self):
        tab = QWidget()
        tab_layout = QVBoxLayout(tab)
        tab_layout.setContentsMargins(10, 10, 10, 10)
        tab_layout.setSpacing(10)

        hint = QLabel(
            "Đổi tên nhanh layer theo danh mục chuẩn. "
            "Chọn tên cũ ở bên trái, rồi chọn tên mới phù hợp (_P/_L/_A) ở bên phải."
        )
        hint.setObjectName("sectionHint")
        hint.setWordWrap(True)
        tab_layout.addWidget(hint)

        box = QGroupBox("Đổi tên nhanh")
        box_layout = QVBoxLayout(box)

        split = QSplitter(ORIENTATION_HORIZONTAL)
        split.setObjectName("hosogisWorkSplit")
        split.setChildrenCollapsible(False)
        split.setHandleWidth(5)
        old_box = QGroupBox("Tên cũ (layer hiện tại)")
        old_layout = QVBoxLayout(old_box)
        self.rename_old_list = QListWidget()
        self.rename_old_list.setMinimumHeight(180)
        self.rename_old_list.currentItemChanged.connect(self.on_rename_old_layer_changed)
        old_layout.addWidget(self.rename_old_list)

        new_box = QGroupBox("Tên mới (chọn từ danh mục)")
        new_layout = QVBoxLayout(new_box)
        self.rename_new_list = QListWidget()
        self.rename_new_list.setMinimumHeight(180)
        new_layout.addWidget(self.rename_new_list)

        split.addWidget(old_box)
        split.addWidget(new_box)
        split.setSizes([400, 400])
        box_layout.addWidget(split, 1)

        btn_row = QHBoxLayout()
        btn_refresh = QPushButton("Làm mới danh sách")
        btn_refresh.setObjectName("ghost")
        btn_apply = QPushButton("Đổi tên layer đã chọn")
        btn_refresh.clicked.connect(self.refresh_rename_layers)
        btn_apply.clicked.connect(self.apply_selected_rename)
        btn_row.addWidget(btn_refresh)
        btn_row.addWidget(btn_apply)
        box_layout.addLayout(btn_row)

        tab_layout.addWidget(box, 1)
        return tab

    def _build_tab_export(self):
        tab = QWidget()
        tab_layout = QVBoxLayout(tab)
        tab_layout.setContentsMargins(10, 10, 10, 10)
        tab_layout.setSpacing(10)

        hint = QLabel("Xuất/nhập dữ liệu GDB, GPKG theo cấu trúc nhóm nghiệp vụ.")
        hint.setObjectName("sectionHint")
        hint.setWordWrap(True)
        tab_layout.addWidget(hint)

        import_box = QGroupBox("Nhập dữ liệu GDB")
        import_layout = QVBoxLayout(import_box)
        import_layout.setContentsMargins(12, 14, 12, 12)
        import_layout.setSpacing(10)
        import_intro = QLabel(
            "Chọn thư mục File Geodatabase (*.gdb) trên máy, sau đó nhập các lớp vào dự án QGIS."
        )
        import_intro.setWordWrap(True)
        import_intro.setObjectName("sectionHint")
        import_layout.addWidget(import_intro)

        import_row = QHBoxLayout()
        import_row.setSpacing(8)
        self.input_gdb_path = QLineEdit()
        self.input_gdb_path.setPlaceholderText("Chọn thư mục File Geodatabase (*.gdb)")
        btn_browse_gdb = QPushButton("Chọn GDB")
        btn_browse_gdb.setObjectName("ghost")
        btn_browse_gdb.clicked.connect(self.choose_gdb_folder)
        import_row.addWidget(self.input_gdb_path, 1)
        import_row.addWidget(btn_browse_gdb)
        import_layout.addLayout(import_row)

        btn_import_gdb = QPushButton("Chạy nhập GDB")
        btn_import_gdb.setMinimumHeight(36)
        btn_import_gdb.clicked.connect(self.import_from_gdb)
        import_layout.addWidget(btn_import_gdb)

        export_box = QGroupBox("Xuất dữ liệu")
        export_layout = QVBoxLayout(export_box)
        export_layout.setContentsMargins(12, 14, 12, 12)
        export_layout.setSpacing(10)
        export_intro = QLabel(
            "Xuất các lớp vector trong dự án hiện tại sang định dạng GDB hoặc GPKG."
        )
        export_intro.setWordWrap(True)
        export_intro.setObjectName("sectionHint")
        export_layout.addWidget(export_intro)

        btn_row_export = QHBoxLayout()
        btn_row_export.setSpacing(8)
        btn_run_gdb = QPushButton("Xuất GDB")
        btn_run_gdb.clicked.connect(self.export_to_gdb)
        btn_run_gpkg = QPushButton("Xuất GPKG")
        btn_run_gpkg.clicked.connect(self.export_to_gpkg)
        for b in (btn_run_gdb, btn_run_gpkg):
            b.setMinimumHeight(36)
        btn_row_export.addWidget(btn_run_gdb, 1)
        btn_row_export.addWidget(btn_run_gpkg, 1)
        export_layout.addLayout(btn_row_export)

        tab_layout.addWidget(import_box, 0)
        tab_layout.addWidget(export_box, 0)
        tab_layout.addStretch(1)
        return tab

    def choose_gdb_folder(self):
        gdb_path = QFileDialog.getExistingDirectory(
            self,
            "Chọn thư mục File Geodatabase (.gdb)",
            "",
        )
        if gdb_path:
            self.input_gdb_path.setText(gdb_path)

    def _add_labeled_input(self, layout, text, widget):
        if hasattr(layout, "addRow"):
            layout.addRow(f"{text}:", widget)
        else:
            layout.addWidget(QLabel(text))
            layout.addWidget(widget)

    def log(self, message):
        text = str(message)
        lower_text = text.lower()

        color = "#1f2937"  # mặc định
        if "lỗi" in lower_text or "không thành công" in lower_text or "thất bại" in lower_text:
            color = "#b91c1c"
        elif "hoàn tất" in lower_text or text.strip().startswith("  +"):
            color = "#047857"
        elif "đang" in lower_text or "bắt đầu" in lower_text:
            color = "#1d4ed8"
        elif "hủy" in lower_text:
            color = "#b45309"

        safe_text = (
            text.replace("&", "&amp;")
            .replace("<", "&lt;")
            .replace(">", "&gt;")
        )
        self.log_edit.append(f'<span style="color:{color};">{safe_text}</span>')
        self.log_edit.moveCursor(TEXT_CURSOR_END)
        QApplication.processEvents()

    def show_done_message(self):
        QMessageBox.information(
            self,
            "Hoàn tất",
            f"Quá trình đã hoàn tất.\n\nHoSoGIS · Phát triển & bản quyền © {_COPYRIGHT_YEAR} {_DEV_BRAND}.\nCảm ơn đã sử dụng.",
        )

    def _question_yes_no(self, title, text, default_no=True):
        try:
            yes = QMessageBox.StandardButton.Yes
            no = QMessageBox.StandardButton.No
            mask = yes | no
            default_btn = no if default_no else yes
            reply = QMessageBox.question(self, title, text, mask, default_btn)
            return reply == yes
        except AttributeError:
            reply = QMessageBox.question(
                self,
                title,
                text,
                QMessageBox.Yes | QMessageBox.No,
                QMessageBox.No if default_no else QMessageBox.Yes,
            )
            return reply == QMessageBox.Yes

    def choose_cad_file(self):
        file_path, _ = QFileDialog.getOpenFileName(
            self,
            "Chọn file CAD (Nên dùng DXF)",
            "",
            "DXF Files (*.dxf);;CAD Files (*.dxf *.dwg)",
        )
        if file_path:
            self.input_cad_path.setText(file_path)

    def refresh_vector_layers(self):
        self.list_layers.clear()
        all_layers = QgsProject.instance().mapLayers().values()
        vector_layers = [lyr for lyr in all_layers if lyr.type() == QgsMapLayerType.VectorLayer]
        for layer in vector_layers:
            item = QListWidgetItem(layer.name())
            item.setFlags(item.flags() | FLAG_CHECKABLE)
            item.setCheckState(CHK_CHECKED)
            item.setData(USER_ROLE, layer.id())
            self.list_layers.addItem(item)
        self.filter_attribute_layers()
        self.update_selected_layer_count()
        if self.list_layers.count() > 0:
            self.list_layers.setCurrentRow(0)
        else:
            self._clear_attribute_preview()
        self.refresh_rename_layers()
        self.log(f"Đã tải danh sách layer: {len(vector_layers)} layer vector.")

    def select_all_layers(self):
        for i in range(self.list_layers.count()):
            self.list_layers.item(i).setCheckState(CHK_CHECKED)
        self.update_selected_layer_count()

    def unselect_all_layers(self):
        for i in range(self.list_layers.count()):
            self.list_layers.item(i).setCheckState(CHK_UNCHECKED)
        self.update_selected_layer_count()

    def filter_attribute_layers(self):
        if not hasattr(self, "list_layers"):
            return
        keyword = ""
        if hasattr(self, "input_layer_filter"):
            keyword = self.input_layer_filter.text().strip().lower()
        for i in range(self.list_layers.count()):
            item = self.list_layers.item(i)
            item.setHidden(keyword not in item.text().lower())

    def update_selected_layer_count(self):
        if not hasattr(self, "lbl_layer_count") or not hasattr(self, "list_layers"):
            return
        selected_count = 0
        visible_count = 0
        for i in range(self.list_layers.count()):
            item = self.list_layers.item(i)
            if not item.isHidden():
                visible_count += 1
            if item.checkState() == CHK_CHECKED:
                selected_count += 1
        self.lbl_layer_count.setText(f"Đã chọn: {selected_count} | Hiển thị: {visible_count}")

    def _selected_vector_layers(self):
        id_map = QgsProject.instance().mapLayers()
        selected = []
        for i in range(self.list_layers.count()):
            item = self.list_layers.item(i)
            if item.checkState() == CHK_CHECKED:
                layer_id = item.data(USER_ROLE)
                layer = id_map.get(layer_id)
                if layer and layer.type() == QgsMapLayerType.VectorLayer:
                    selected.append(layer)
        return selected

    def _clear_attribute_preview(self):
        if not hasattr(self, "attr_preview_table"):
            return
        self.attr_preview_table.setRowCount(0)
        if hasattr(self, "attr_preview_meta"):
            self.attr_preview_meta.setText("Chọn 1 layer để xem cấu trúc trường dữ liệu.")

    def on_attribute_layer_changed(self, current, previous):
        del previous  # tránh cảnh báo biến không dùng
        if not hasattr(self, "attr_preview_table"):
            return
        self._clear_attribute_preview()
        if current is None:
            return

        layer_id = current.data(USER_ROLE)
        layer = QgsProject.instance().mapLayers().get(layer_id)
        if not layer or layer.type() != QgsMapLayerType.VectorLayer:
            return
        self._fill_attribute_preview_table(layer)

    def _preview_target_layer(self):
        item = self.list_layers.currentItem()
        if item is None or item.isHidden():
            return None
        layer_id = item.data(USER_ROLE)
        layer = QgsProject.instance().mapLayers().get(layer_id)
        if layer and layer.type() == QgsMapLayerType.VectorLayer:
            return layer
        return None

    def _sanitize_field_name(self, value):
        text = str(value).strip()
        if not text:
            return ""
        for ch in '<>:"/\\|?*':
            text = text.replace(ch, "_")
        text = text.replace(" ", "_")
        return text

    def _fill_attribute_preview_table(self, layer):
        if not hasattr(self, "attr_preview_table"):
            return
        fields = layer.fields()
        if hasattr(self, "attr_preview_meta"):
            self.attr_preview_meta.setText(
                f"Layer: «{layer.name()}» — Tổng số trường: {fields.count()}"
            )
        self.attr_preview_table.setRowCount(fields.count())
        for i in range(fields.count()):
            fld = fields.at(i)
            name_item = QTableWidgetItem(fld.name())
            name_item.setData(USER_ROLE, i)
            type_item = QTableWidgetItem(fld.typeName() or str(fld.type()))
            length_val = fld.length()
            length_text = str(length_val) if int(length_val) > 0 else "-"
            length_item = QTableWidgetItem(length_text)
            self.attr_preview_table.setItem(i, 0, name_item)
            self.attr_preview_table.setItem(i, 1, type_item)
            self.attr_preview_table.setItem(i, 2, length_item)

        header = self.attr_preview_table.horizontalHeader()
        header.setStretchLastSection(True)
        header.setSectionResizeMode(0, HEADER_STRETCH)
        header.setSectionResizeMode(1, HEADER_STRETCH)

    def add_attribute_field_manual(self):
        layer = self._preview_target_layer()
        if layer is None:
            self.log("Chưa chọn layer để thêm thuộc tính.")
            return
        pr = layer.dataProvider()
        if pr is None:
            self.log("Layer không có nhà cung cấp dữ liệu, không thể thêm trường.")
            return
        caps = pr.capabilities()
        add_cap = getattr(QgsVectorDataProvider, "AddAttributes", None)
        if add_cap is not None and not (caps & add_cap):
            QMessageBox.warning(
                self,
                "Không thể thêm thuộc tính",
                "Nguồn dữ liệu của layer này không cho phép thêm cột mới.",
            )
            return

        dlg = QDialog(self)
        dlg.setObjectName("hosogisSubDialog")
        dlg.setWindowTitle("Thêm thuộc tính")
        dlg.setMinimumWidth(400)
        outer = QVBoxLayout(dlg)
        outer.setContentsMargins(16, 16, 16, 16)
        outer.setSpacing(12)
        form = QFormLayout()
        form.setLabelAlignment(ALIGN_VCENTER)
        form.setHorizontalSpacing(10)
        form.setVerticalSpacing(10)
        le_name = QLineEdit()
        le_name.setPlaceholderText("Ví dụ: tenDoiTuong")
        combo_type = QComboBox()
        combo_type.addItem("Văn bản (String)", "string")
        combo_type.addItem("Số nguyên (Integer)", "int")
        combo_type.addItem("Số thực (Double)", "double")
        combo_type.addItem("Đúng/Sai (Boolean)", "bool")
        combo_type.addItem("Ngày (Date)", "date")
        spin_len = QSpinBox()
        spin_len.setButtonSymbols(SPIN_NO_BUTTONS)
        spin_len.setRange(1, 255)
        spin_len.setValue(80)
        spin_len.setEnabled(True)

        def on_type_changed(_idx):
            spin_len.setEnabled(combo_type.currentData() == "string")

        combo_type.currentIndexChanged.connect(on_type_changed)
        form.addRow("Tên trường:", le_name)
        form.addRow("Kiểu dữ liệu:", combo_type)
        form.addRow("Độ dài (chuỗi):", spin_len)
        outer.addLayout(form)
        dlg_btn_row = QHBoxLayout()
        dlg_btn_row.setSpacing(8)
        dlg_btn_row.addStretch(1)
        btn_dlg_cancel = QPushButton("Hủy")
        btn_dlg_cancel.setObjectName("ghost")
        btn_dlg_ok = QPushButton("Thêm trường")
        btn_dlg_cancel.clicked.connect(dlg.reject)
        btn_dlg_ok.clicked.connect(dlg.accept)
        dlg_btn_row.addWidget(btn_dlg_cancel)
        dlg_btn_row.addWidget(btn_dlg_ok)
        outer.addLayout(dlg_btn_row)

        exec_res = dlg.exec()
        accepted_code = getattr(QDialog, "Accepted", 1)
        dialog_ok = exec_res == accepted_code
        dialog_code = getattr(QDialog, "DialogCode", None)
        if dialog_code is not None and hasattr(dialog_code, "Accepted"):
            dialog_ok = dialog_ok or exec_res == dialog_code.Accepted
        if not dialog_ok:
            return

        raw_name = self._sanitize_field_name(le_name.text())
        if not raw_name:
            self.log("Tên trường không hợp lệ.")
            return
        if layer.fields().indexOf(raw_name) >= 0:
            self.log(f"Trường «{raw_name}» đã tồn tại trên layer.")
            return

        vtype = combo_type.currentData()
        if vtype == "string":
            new_field = QgsField(raw_name, QVariant.String, len=spin_len.value())
        elif vtype == "int":
            new_field = QgsField(raw_name, QVariant.Int)
        elif vtype == "double":
            new_field = QgsField(raw_name, QVariant.Double, len=20, prec=8)
        elif vtype == "bool":
            new_field = QgsField(raw_name, QVariant.Bool)
        elif vtype == "date":
            new_field = QgsField(raw_name, QVariant.Date)
        else:
            new_field = QgsField(raw_name, QVariant.String, len=80)

        layer.startEditing()
        pr = layer.dataProvider()
        _added = pr.addAttributes([new_field])
        if _added is False:
            layer.rollBack()
            self.log(f"Lỗi: không thêm được trường «{raw_name}».")
            return
        layer.updateFields()
        if not layer.commitChanges():
            layer.rollBack()
            self.log("Lỗi: không lưu được thay đổi cấu trúc trường (đã hủy).")
            return
        self.log(f"Đã thêm trường «{raw_name}» vào layer «{layer.name()}».")
        self._fill_attribute_preview_table(layer)

    def rename_attribute_field_manual(self):
        layer = self._preview_target_layer()
        if layer is None:
            self.log("Chưa chọn layer để đổi tên thuộc tính.")
            return
        pr = layer.dataProvider()
        if pr is None:
            return
        caps = pr.capabilities()
        ren_cap = getattr(QgsVectorDataProvider, "RenameAttributes", None)
        if ren_cap is not None and not (caps & ren_cap):
            QMessageBox.warning(
                self,
                "Không thể đổi tên",
                "Nguồn dữ liệu của layer này không hỗ trợ đổi tên cột.",
            )
            return
        if not hasattr(layer, "renameAttribute"):
            QMessageBox.warning(
                self,
                "Không thể đổi tên",
                "Phiên bản QGIS hiện tại không hỗ trợ đổi tên trường theo cách này.",
            )
            return

        rows = self.attr_preview_table.selectionModel().selectedRows()
        if len(rows) != 1:
            self.log("Hãy chọn đúng một dòng trong bảng thuộc tính để đổi tên.")
            return
        row = rows[0].row()
        name_item = self.attr_preview_table.item(row, 0)
        if name_item is None:
            return
        fidx = name_item.data(USER_ROLE)
        if fidx is None or int(fidx) < 0:
            return
        old_name = layer.fields().at(int(fidx)).name()
        new_name, ok = QInputDialog.getText(
            self,
            "Đổi tên thuộc tính",
            f"Tên mới cho trường «{old_name}»:",
            text=old_name,
        )
        if not ok:
            return
        new_name = self._sanitize_field_name(new_name)
        if not new_name:
            self.log("Tên mới không hợp lệ.")
            return
        if new_name == old_name:
            return
        if layer.fields().indexOf(new_name) >= 0:
            self.log(f"Tên «{new_name}» đã tồn tại trên layer.")
            return

        layer.startEditing()
        _ren = layer.renameAttribute(int(fidx), new_name)
        if _ren is False:
            layer.rollBack()
            self.log(f"Lỗi: không đổi tên được trường «{old_name}».")
            return
        if not layer.commitChanges():
            layer.rollBack()
            self.log("Lỗi: không lưu được đổi tên trường (đã hủy).")
            return
        self.log(f"Đã đổi tên trường: «{old_name}» → «{new_name}» (layer «{layer.name()}»).")
        self._fill_attribute_preview_table(layer)

    def delete_attribute_fields_manual(self):
        layer = self._preview_target_layer()
        if layer is None:
            self.log("Chưa chọn layer để xóa thuộc tính.")
            return
        pr = layer.dataProvider()
        if pr is None:
            return
        caps = pr.capabilities()
        del_cap = getattr(QgsVectorDataProvider, "DeleteAttributes", None)
        if del_cap is not None and not (caps & del_cap):
            QMessageBox.warning(
                self,
                "Không thể xóa thuộc tính",
                "Nguồn dữ liệu của layer này không cho phép xóa cột.",
            )
            return

        rows = self.attr_preview_table.selectionModel().selectedRows()
        if not rows:
            self.log("Hãy chọn ít nhất một dòng trong bảng thuộc tính để xóa.")
            return
        indices = []
        for mi in rows:
            name_item = self.attr_preview_table.item(mi.row(), 0)
            if name_item is None:
                continue
            fidx = name_item.data(USER_ROLE)
            if fidx is not None:
                indices.append(int(fidx))
        if not indices:
            return
        indices = sorted(set(indices), reverse=True)
        names = [layer.fields().at(i).name() for i in indices]
        preview = ", ".join(names[:8])
        if len(names) > 8:
            preview += ", …"
        if not self._question_yes_no(
            "Xác nhận xóa thuộc tính",
            f"Sẽ xóa {len(names)} trường khỏi layer «{layer.name()}»:\n\n{preview}\n\n"
            "Thao tác này không thể hoàn tác từ HoSoGIS. Tiếp tục?",
            default_no=True,
        ):
            self.log("Đã hủy xóa thuộc tính.")
            return

        layer.startEditing()
        pr = layer.dataProvider()
        _deleted = pr.deleteAttributes(indices)
        if _deleted is False:
            layer.rollBack()
            self.log("Lỗi: không xóa được các trường đã chọn.")
            return
        layer.updateFields()
        if not layer.commitChanges():
            layer.rollBack()
            self.log("Lỗi: không lưu được sau khi xóa trường (đã hủy).")
            return
        self.log(f"Đã xóa {len(names)} trường khỏi layer «{layer.name()}».")
        self._fill_attribute_preview_table(layer)

    def _geometry_suffix_for_layer(self, layer):
        if layer.geometryType() == QgsWkbTypes.PointGeometry:
            return "_P"
        if layer.geometryType() == QgsWkbTypes.LineGeometry:
            return "_L"
        if layer.geometryType() == QgsWkbTypes.PolygonGeometry:
            return "_A"
        return ""

    def refresh_rename_layers(self):
        if not hasattr(self, "rename_old_list"):
            return
        self.rename_old_list.clear()
        self.rename_new_list.clear()
        all_layers = QgsProject.instance().mapLayers().values()
        vector_layers = [lyr for lyr in all_layers if lyr.type() == QgsMapLayerType.VectorLayer]
        for layer in vector_layers:
            item = QListWidgetItem(layer.name())
            item.setData(USER_ROLE, layer.id())
            self.rename_old_list.addItem(item)
        if self.rename_old_list.count() > 0:
            self.rename_old_list.setCurrentRow(0)

    def on_rename_old_layer_changed(self, current, previous):
        del previous  # tránh cảnh báo biến không dùng
        if not hasattr(self, "rename_new_list"):
            return
        self.rename_new_list.clear()
        if current is None:
            return

        layer_id = current.data(USER_ROLE)
        layer = QgsProject.instance().mapLayers().get(layer_id)
        if not layer:
            return
        suffix = self._geometry_suffix_for_layer(layer)
        if not suffix:
            return

        options = [name for name in _QUICK_LAYER_NAME_OPTIONS if name.endswith(suffix)]
        for name in options:
            self.rename_new_list.addItem(QListWidgetItem(name))

    def apply_selected_rename(self):
        if not hasattr(self, "rename_old_list") or not hasattr(self, "rename_new_list"):
            return
        old_item = self.rename_old_list.currentItem()
        new_item = self.rename_new_list.currentItem()
        if old_item is None:
            self.log("Chưa chọn layer ở cột Tên cũ.")
            return
        if new_item is None:
            self.log("Chưa chọn tên mới ở cột Tên mới.")
            return

        layer_id = old_item.data(USER_ROLE)
        layer = QgsProject.instance().mapLayers().get(layer_id)
        if not layer:
            self.log("Layer đã bị xóa khỏi dự án, vui lòng làm mới danh sách.")
            return

        new_name = new_item.text().strip()
        if not new_name:
            self.log("Tên mới không hợp lệ.")
            return
        old_name = layer.name()
        if old_name == new_name:
            self.log(f"Layer «{old_name}» đã có sẵn tên này.")
            return

        # Cảnh báo khi tên mới đã tồn tại ở layer khác trong project
        duplicated_layers = []
        for lyr in QgsProject.instance().mapLayers().values():
            if lyr.id() == layer.id():
                continue
            if lyr.name() == new_name:
                duplicated_layers.append(lyr)
        if duplicated_layers:
            confirm = self._question_yes_no(
                "Cảnh báo trùng tên layer",
                "Tên mới bạn chọn đang trùng với layer đã tồn tại:\n\n"
                f"«{new_name}»\n\n"
                "Việc trùng tên có thể gây khó phân biệt khi thao tác/xuất dữ liệu.\n"
                "Bạn vẫn muốn tiếp tục đổi tên?",
                default_no=True,
            )
            if not confirm:
                self.log(f"Đã hủy đổi tên do trùng tên layer: «{new_name}».")
                return

        layer.setName(new_name)
        old_item.setText(new_name)
        for i in range(self.list_layers.count()):
            item_attr = self.list_layers.item(i)
            if item_attr.data(USER_ROLE) == layer.id():
                item_attr.setText(new_name)
                break
        self.log(f"Đã đổi tên layer: «{old_name}» -> «{new_name}».")

    def export_attributes_excel(self):
        selected = self._selected_vector_layers()
        if not selected:
            self.log("Bạn chưa chọn lớp vector nào.")
            return
        home = os.path.expanduser("~")
        if len(selected) == 1:
            default_name = sanitize_filename(selected[0].name()) + ".xlsx"
            path, _ = QFileDialog.getSaveFileName(
                self,
                "Lưu thuộc tính ra Excel",
                os.path.join(home, default_name),
                "Excel (*.xlsx)",
            )
        else:
            n = len(selected)
            default_zip = sanitize_filename(selected[0].name()) + f"_{n}lop.zip"
            path, _ = QFileDialog.getSaveFileName(
                self,
                "Lưu thuộc tính (ZIP chứa nhiều file Excel)",
                os.path.join(home, default_zip),
                "ZIP (*.zip)",
            )
        if not path:
            self.log("Đã hủy xuất Excel.")
            return
        self.log(f"Đang xuất thuộc tính của {len(selected)} lớp…")
        out_path, err = export_layers_attributes_excel(selected, path)
        if err:
            self.log(f"Lỗi xuất Excel: {err}")
            return
        self.log(f"Đã lưu: {out_path}")
        self.show_done_message()

    def import_attributes_excel(self):
        selected = self._selected_vector_layers()
        if not selected:
            self.log("Bạn chưa chọn lớp vector nào.")
            return
        if len(selected) > 1:
            QMessageBox.warning(
                self,
                "Nhập thuộc tính — chỉ một lớp",
                "Nhập dữ liệu từ Excel chỉ thực hiện được cho một lớp mỗi lần.\n\n"
                f"Hiện bạn đang chọn {len(selected)} lớp. Hãy dùng «Bỏ chọn tất cả» rồi chỉ đánh dấu đúng một lớp cần nhập.",
            )
            self.log(f"Đã dừng: nhập Excel cần đúng 1 lớp (đang chọn {len(selected)} lớp).")
            return

        layer = selected[0]
        path, _ = QFileDialog.getOpenFileName(
            self,
            "Chọn file Excel để nhập thuộc tính",
            "",
            "Excel (*.xlsx);;ZIP (*.zip)",
        )
        if not path:
            self.log("Đã hủy nhập Excel.")
            return
        if not self._question_yes_no(
            "Xác nhận nhập từ Excel",
            "Sẽ ghi dữ liệu từ file:\n\n"
            f"«{os.path.basename(path)}»\n\n"
            "vào các đối tượng của lớp (khớp theo cột «qgis_fid»):\n\n"
            f"«{layer.name()}»\n\n"
            "Tiếp tục nhập?",
        ):
            self.log("Đã hủy nhập Excel (chưa xác nhận sau khi chọn file).")
            return
        self.log(f"Đang nhập thuộc tính từ «{os.path.basename(path)}» vào lớp «{layer.name()}»…")
        err = import_layers_attributes_excel([layer], path, log_print=self.log)
        if err:
            self.log(f"Lỗi nhập Excel: {err}")
            return
        self.log("--- HOÀN TẤT nhập thuộc tính từ Excel ---")
        self.show_done_message()

    def _get_field_idx(self, layer, field_name):
        idx = layer.fields().indexOf(field_name)
        if idx == -1:
            idx = layer.fields().indexOf(field_name[:10])
        return idx

    def _sanitize_name(self, value):
        text = str(value).strip()
        if not text:
            return "unnamed"
        invalid_chars = '<>:"/\\|?*'
        for ch in invalid_chars:
            text = text.replace(ch, "_")
        text = text.replace(" ", "_").replace("-", "_")
        return text.strip("._") or "unnamed"

    # --- CHUC NANG 1: NHAP BAN VE ---
    def import_and_split_cad(self):
        file_path = self.input_cad_path.text().strip()
        if not file_path:
            file_path, _ = QFileDialog.getOpenFileName(
                self,
                "Chọn file CAD (Nên dùng DXF)",
                "",
                "DXF Files (*.dxf);;CAD Files (*.dxf *.dwg)",
            )
            if not file_path:
                self.log("Đã hủy chọn file CAD.")
                return
            self.input_cad_path.setText(file_path)

        crs_dialog = QgsProjectionSelectionDialog()
        crs_dialog.setWindowTitle("Chọn hệ tọa độ cho bản vẽ CAD")
        if not crs_dialog.exec():
            self.log("Đã hủy chọn hệ tọa độ.")
            return
        selected_crs = crs_dialog.crs()

        file_name = os.path.basename(file_path)
        file_base_name = os.path.splitext(file_name)[0]
        self.log(f"Đang xử lý file: {file_name} với hệ tọa độ {selected_crs.authid()}...")

        dxf_full_layer_names = _extract_dxf_layer_names(file_path) if file_path.lower().endswith(".dxf") else []

        root = QgsProject.instance().layerTreeRoot()
        cad_group = root.addGroup(f"CAD_{file_base_name}")

        geometry_mapping = {
            "Point": ("", "Điểm (Point)"),
            "LineString": ("", "Đường (Line)"),
            "Polygon": ("", "Vùng (Polygon)"),
        }
        # Hậu tố tên lớp theo loại hình học (bước 1 — nhập CAD)
        _cad_geom_suffix = {"Point": "_P", "LineString": "_L", "Polygon": "_A"}

        layer_count = 0
        subgroups = {}

        for geom_type, (suffix, group_name) in geometry_mapping.items():
            uri = f"{file_path}|layername=entities|geometrytype={geom_type}"
            temp_layer = QgsVectorLayer(uri, "temp", "ogr")

            if not temp_layer.isValid() or temp_layer.featureCount() == 0:
                continue

            idx = temp_layer.fields().indexOf("Layer")
            if idx == -1:
                continue

            unique_cad_layers = temp_layer.uniqueValues(idx)

            for cad_layer in unique_cad_layers:
                if not cad_layer:
                    continue

                cad_layer_str = str(cad_layer)
                restored_layer = _restore_layer_name(cad_layer_str, dxf_full_layer_names)

                clean_name = restored_layer.strip().replace(" ", "_").replace("-", "_")
                layer_name = clean_name + _cad_geom_suffix.get(geom_type, "")
                new_layer = QgsVectorLayer(uri, layer_name, "ogr")
                safe_cad_layer = cad_layer_str.replace("'", "''")
                new_layer.setSubsetString(f"\"Layer\" = '{safe_cad_layer}'")

                if new_layer.isValid() and new_layer.featureCount() > 0:
                    editable_layer = new_layer.materialize(QgsFeatureRequest())
                    editable_layer.setName(layer_name)
                    editable_layer.setCrs(selected_crs)
                    QgsProject.instance().addMapLayer(editable_layer, False)

                    if group_name not in subgroups:
                        subgroups[group_name] = cad_group.addGroup(group_name)

                    subgroups[group_name].addLayer(editable_layer)
                    layer_count += 1
                    self.log(f"  + Đã tách và mở khóa: {layer_name} -> Nhóm: {group_name}")

        if layer_count > 0:
            self.log(f"HOÀN TẤT! Đã import và phân loại {layer_count} layer.")
            self.refresh_vector_layers()
            self.show_done_message()
        else:
            self.log("KHÔNG THÀNH CÔNG! Không tìm thấy dữ liệu hợp lệ (thử Save As DWG -> DXF).")

    # --- CHUC NANG 2: TAO THUOC TINH ---
    def add_fields_and_data(self):
        selected_layers = self._selected_vector_layers()
        if not selected_layers:
            self.log("Bạn chưa chọn layer nào để xử lý.")
            return

        ma_tt_qh = self.input_ma_tt.text().strip()
        ma_hs_qh = self.input_ma_hs.text().strip()
        ma_dt_goc = self.input_ma_dt.text().strip()
        ten_dt = self.input_ten_dt.text().strip()
        phan_loai = self.input_phan_loai.text().strip()
        ghi_chu = self.input_ghi_chu.text().strip()
        delete_old_fields = self.chk_delete_old.isChecked()

        self.log(f"Bắt đầu xử lý {len(selected_layers)} layer đã chọn...")

        fields_to_add = [
            QgsField("maThongTinQH", QVariant.String, len=15),
            QgsField("maHoSoQH", QVariant.String, len=15),
            QgsField("maDoiTuong", QVariant.String, len=100),
            QgsField("tenDoiTuong", QVariant.String, len=100),
            QgsField("phanLoai", QVariant.String, len=250),
            QgsField("ghiChu", QVariant.String, len=250),
        ]

        for layer in selected_layers:
            self.log(f"Đang cập nhật: {layer.name()}")
            layer.startEditing()
            pr = layer.dataProvider()

            if delete_old_fields:
                old_field_count = len(layer.fields())
                if old_field_count > 0:
                    pr.deleteAttributes(list(range(old_field_count)))
                    layer.updateFields()
                    self.log(f"  + Đã xóa {old_field_count} thuộc tính cũ")

            existing_fields = layer.fields().names()
            new_fields = [
                f
                for f in fields_to_add
                if f.name() not in existing_fields and f.name()[:10] not in existing_fields
            ]
            if new_fields:
                pr.addAttributes(new_fields)
                layer.updateFields()

            idx_maThongTinQH = self._get_field_idx(layer, "maThongTinQH")
            idx_maHoSoQH = self._get_field_idx(layer, "maHoSoQH")
            idx_maDoiTuong = self._get_field_idx(layer, "maDoiTuong")
            idx_tenDoiTuong = self._get_field_idx(layer, "tenDoiTuong")
            idx_phanLoai = self._get_field_idx(layer, "phanLoai")
            idx_ghiChu = self._get_field_idx(layer, "ghiChu")

            update_dict = {}
            for feat in layer.getFeatures():
                attr_map = {}
                if idx_maThongTinQH != -1:
                    attr_map[idx_maThongTinQH] = ma_tt_qh
                if idx_maHoSoQH != -1:
                    attr_map[idx_maHoSoQH] = ma_hs_qh
                if idx_maDoiTuong != -1:
                    attr_map[idx_maDoiTuong] = ma_dt_goc
                if idx_tenDoiTuong != -1 and ten_dt:
                    attr_map[idx_tenDoiTuong] = ten_dt
                if idx_phanLoai != -1 and phan_loai:
                    attr_map[idx_phanLoai] = phan_loai
                if idx_ghiChu != -1 and ghi_chu:
                    attr_map[idx_ghiChu] = ghi_chu
                update_dict[feat.id()] = attr_map

            pr.changeAttributeValues(update_dict)
            layer.commitChanges()
            self.log(f"  + Hoàn tất: {layer.name()}")

        self.log("--- ĐÃ CHẠY XONG TOÀN BỘ YÊU CẦU ---")
        self.show_done_message()

    # --- CHUC NANG 3: TAO GDB ---
    def export_to_gdb(self):
        gdb_path, _ = QFileDialog.getSaveFileName(
            self,
            "Chọn nơi lưu và đặt tên file Geodatabase",
            "",
            "Esri File Geodatabase (*.gdb)",
        )

        if not gdb_path:
            self.log("Đã hủy thao tác lưu file.")
            return

        if not gdb_path.endswith(".gdb"):
            gdb_path += ".gdb"

        self.log(f"Bắt đầu xuất dữ liệu ra: {gdb_path}")

        gdb_exists = os.path.exists(gdb_path)
        exported_layers = 0

        def process_group(group):
            nonlocal gdb_exists, exported_layers
            group_name = group.name()
            if group_name == "base map":
                return

            self.log(f"Đang quét nhóm: {group_name}")
            for child in group.children():
                if isinstance(child, QgsLayerTreeGroup):
                    process_group(child)
                elif isinstance(child, QgsLayerTreeLayer):
                    layer = child.layer()
                    if layer and layer.type() == QgsMapLayerType.VectorLayer:
                        safe_layer_name = self._sanitize_name(layer.name())
                        options = QgsVectorFileWriter.SaveVectorOptions()
                        options.driverName = "OpenFileGDB"
                        options.layerName = safe_layer_name
                        if not gdb_exists:
                            options.actionOnExistingFile = QgsVectorFileWriter.CreateOrOverwriteFile
                        else:
                            options.actionOnExistingFile = QgsVectorFileWriter.CreateOrOverwriteLayer
                        options.layerOptions = [f"FEATURE_DATASET={self._sanitize_name(group_name)}"]

                        error = QgsVectorFileWriter.writeAsVectorFormatV3(
                            layer,
                            gdb_path,
                            QgsProject.instance().transformContext(),
                            options,
                        )
                        if error[0] == QgsVectorFileWriter.NoError:
                            self.log(f"  + Đã xuất: {layer.name()} -> Nhóm GDB: {group_name}")
                            gdb_exists = True
                            exported_layers += 1
                        else:
                            self.log(f"  - Lỗi khi xuất {layer.name()}: {error}")

        root = QgsProject.instance().layerTreeRoot()
        for child in root.children():
            if isinstance(child, QgsLayerTreeGroup):
                process_group(child)

        if exported_layers > 0:
            self.log(f"--- HOÀN TẤT --- Đã xuất {exported_layers} layer.")
            self.show_done_message()
        else:
            self.log("--- HOÀN TẤT --- Không có layer nào được xuất.")

    def export_to_gpkg(self):
        base_dir = QFileDialog.getExistingDirectory(
            self,
            "Chọn thư mục lưu dữ liệu GPKG",
            "",
        )
        if not base_dir:
            self.log("Đã hủy chọn thư mục lưu.")
            return

        folder_name, ok = QInputDialog.getText(
            self,
            "Tên thư mục tổng",
            "Nhập tên thư mục tổng để chứa dữ liệu xuất GPKG:",
            text="export_gpkg",
        )
        if not ok:
            self.log("Đã hủy nhập tên thư mục tổng.")
            return

        folder_name = self._sanitize_name(folder_name)
        output_root = os.path.join(base_dir, folder_name)
        os.makedirs(output_root, exist_ok=True)
        self.log(f"Bắt đầu xuất GPKG vào thư mục: {output_root}")

        exported_layers = 0

        def process_group(group):
            nonlocal exported_layers
            group_name = group.name()
            if group_name.lower() == "base map":
                return

            safe_group_name = self._sanitize_name(group_name)
            group_dir = os.path.join(output_root, safe_group_name)
            os.makedirs(group_dir, exist_ok=True)
            self.log(f"Đang quét nhóm: {group_name}")

            for child in group.children():
                if isinstance(child, QgsLayerTreeGroup):
                    process_group(child)
                elif isinstance(child, QgsLayerTreeLayer):
                    layer = child.layer()
                    if layer and layer.type() == QgsMapLayerType.VectorLayer:
                        safe_layer_name = self._sanitize_name(layer.name())
                        gpkg_path = os.path.join(group_dir, f"{safe_layer_name}.gpkg")

                        options = QgsVectorFileWriter.SaveVectorOptions()
                        options.driverName = "GPKG"
                        options.layerName = safe_layer_name
                        options.actionOnExistingFile = QgsVectorFileWriter.CreateOrOverwriteFile

                        error = QgsVectorFileWriter.writeAsVectorFormatV3(
                            layer,
                            gpkg_path,
                            QgsProject.instance().transformContext(),
                            options,
                        )
                        if error[0] == QgsVectorFileWriter.NoError:
                            self.log(f"  + Đã xuất: {layer.name()} -> {gpkg_path}")
                            exported_layers += 1
                        else:
                            self.log(f"  - Lỗi khi xuất {layer.name()}: {error}")

        root = QgsProject.instance().layerTreeRoot()
        for child in root.children():
            if isinstance(child, QgsLayerTreeGroup):
                process_group(child)

        if exported_layers > 0:
            self.log(f"--- HOÀN TẤT --- Đã xuất {exported_layers} layer ra GPKG.")
            self.show_done_message()
        else:
            self.log("--- HOÀN TẤT --- Không có layer nào được xuất ra GPKG.")

    def _list_gdb_sublayers(self, gdb_path):
        probe = QgsVectorLayer(gdb_path, "__gdb_probe__", "ogr")
        if not probe.isValid() or not probe.dataProvider():
            return []
        sublayers = []
        for raw in probe.dataProvider().subLayers():
            text = str(raw)
            # OGR thường trả dạng: "<idx>!!::!!<layer_name>!!::!!<feature_count>..."
            parts = text.split("!!::!!")
            name = parts[1].strip() if len(parts) > 1 else text.strip()
            if name and name not in sublayers:
                sublayers.append(name)
        return sublayers

    def _geometry_group_name(self, geometry_type):
        if geometry_type == QgsWkbTypes.PointGeometry:
            return "Điểm (Point)"
        if geometry_type == QgsWkbTypes.LineGeometry:
            return "Đường (Line)"
        if geometry_type == QgsWkbTypes.PolygonGeometry:
            return "Vùng (Polygon)"
        return "Khác (Other)"

    def _unique_group_name(self, root, base_name):
        name = base_name
        idx = 2
        while root.findGroup(name) is not None:
            name = f"{base_name}_{idx}"
            idx += 1
        return name

    def import_from_gdb(self):
        gdb_path = self.input_gdb_path.text().strip() if hasattr(self, "input_gdb_path") else ""
        if not gdb_path:
            gdb_path = QFileDialog.getExistingDirectory(
                self,
                "Chọn thư mục File Geodatabase (.gdb)",
                "",
            )
        if not gdb_path:
            self.log("Đã hủy chọn thư mục GDB.")
            return
        if hasattr(self, "input_gdb_path"):
            self.input_gdb_path.setText(gdb_path)
        if not gdb_path.lower().endswith(".gdb"):
            self.log("Đường dẫn đã chọn không phải thư mục .gdb.")
            return

        sublayers = self._list_gdb_sublayers(gdb_path)
        if not sublayers:
            self.log("Không đọc được layer nào từ GDB (hoặc GDB rỗng).")
            return

        root = QgsProject.instance().layerTreeRoot()
        gdb_name = os.path.splitext(os.path.basename(gdb_path))[0]
        group_name = self._unique_group_name(root, f"GDB_{self._sanitize_name(gdb_name)}")
        main_group = root.addGroup(group_name)
        subgroups = {}
        loaded = 0

        self.log(f"Bắt đầu nhập GDB: {gdb_path}")
        for sublayer_name in sublayers:
            uri = f"{gdb_path}|layername={sublayer_name}"
            layer = QgsVectorLayer(uri, sublayer_name, "ogr")
            if not layer.isValid():
                self.log(f"  - Bỏ qua (không hợp lệ): {sublayer_name}")
                continue
            if layer.type() != QgsMapLayerType.VectorLayer:
                self.log(f"  - Bỏ qua (không phải vector): {sublayer_name}")
                continue

            geom_group_name = self._geometry_group_name(layer.geometryType())
            if geom_group_name not in subgroups:
                subgroups[geom_group_name] = main_group.addGroup(geom_group_name)

            QgsProject.instance().addMapLayer(layer, False)
            subgroups[geom_group_name].addLayer(layer)
            loaded += 1
            self.log(f"  + Đã nhập: {sublayer_name} -> Nhóm: {geom_group_name}")

        if loaded > 0:
            self.log(f"HOÀN TẤT! Đã nhập {loaded} layer từ GDB và phân nhóm theo hình học.")
            self.refresh_vector_layers()
            self.show_done_message()
        else:
            self.log("KHÔNG THÀNH CÔNG! Không có layer hợp lệ nào được nhập từ GDB.")


def show_hosogis_gui():
    global HOSOGIS_WINDOW
    try:
        HOSOGIS_WINDOW.close()
    except Exception:
        pass

    HOSOGIS_WINDOW = HoSoGISWindow()
    HOSOGIS_WINDOW.show()
    HOSOGIS_WINDOW.raise_()
    HOSOGIS_WINDOW.activateWindow()


show_hosogis_gui()
