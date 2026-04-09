import os
import tempfile
import zipfile
from collections import defaultdict

from qgis.core import QgsProject, QgsField, QgsMapLayerType
from qgis.PyQt.QtCore import QDate, QDateTime, QVariant, Qt
from qgis.PyQt.QtWidgets import (QDialog, QVBoxLayout, QHBoxLayout, QLabel, 
                                 QLineEdit, QListWidget, QListWidgetItem, 
                                 QPushButton, QDialogButtonBox, QScrollArea, QWidget, QMessageBox, QCheckBox)

# --- XỬ LÝ PHIÊN BẢN PYQT ---
try:
    CHK_CHECKED = Qt.CheckState.Checked
    CHK_UNCHECKED = Qt.CheckState.Unchecked
    FLAG_CHECKABLE = Qt.ItemFlag.ItemIsUserCheckable
    USER_ROLE = Qt.ItemDataRole.UserRole
    BTN_OK = QDialogButtonBox.StandardButton.Ok
    BTN_CANCEL = QDialogButtonBox.StandardButton.Cancel
except AttributeError:
    CHK_CHECKED = Qt.Checked
    CHK_UNCHECKED = Qt.Unchecked
    FLAG_CHECKABLE = Qt.ItemIsUserCheckable
    USER_ROLE = Qt.UserRole
    BTN_OK = QDialogButtonBox.Ok
    BTN_CANCEL = QDialogButtonBox.Cancel

class LayerSelectionDialog(QDialog):
    def __init__(self, vector_layers, parent=None):
        super().__init__(parent)
        self.setWindowTitle("Cập nhật Thuộc Tính Quy Hoạch")
        self.resize(500, 650)
        
        layout = QVBoxLayout(self)

        # Tạo vùng cuộn (Scroll Area) để chứa nhiều ô nhập liệu không bị tràn màn hình
        scroll = QScrollArea(self)
        scroll.setWidgetResizable(True)
        scroll_content = QWidget()
        scroll_layout = QVBoxLayout(scroll_content)

        # 1. Khu vực nhập liệu (Đã bổ sung đầy đủ)
        scroll_layout.addWidget(QLabel("Mã thông tin quy hoạch (maThongTinQH):"))
        self.input_ma_tt = QLineEdit(self)
        scroll_layout.addWidget(self.input_ma_tt)

        scroll_layout.addWidget(QLabel("Mã hồ sơ quy hoạch (maHoSoQH) - VD: 84QHC1000001:"))
        self.input_ma_hs = QLineEdit("84QHC1000001", self)
        scroll_layout.addWidget(self.input_ma_hs)

        scroll_layout.addWidget(QLabel("Mã đối tượng (maDoiTuong) - do người dùng nhập:"))
        self.input_ma_dt = QLineEdit(self)
        scroll_layout.addWidget(self.input_ma_dt)
        
        scroll_layout.addWidget(QLabel("Tên đối tượng (tenDoiTuong) - Bỏ trống nếu nhập sau:"))
        self.input_ten_dt = QLineEdit(self)
        scroll_layout.addWidget(self.input_ten_dt)
        
        scroll_layout.addWidget(QLabel("Phân loại (phanLoai) - Bỏ trống nếu nhập sau:"))
        self.input_phan_loai = QLineEdit(self)
        scroll_layout.addWidget(self.input_phan_loai)
        
        scroll_layout.addWidget(QLabel("Ghi chú (ghiChu) - Bỏ trống nếu nhập sau:"))
        self.input_ghi_chu = QLineEdit(self)
        scroll_layout.addWidget(self.input_ghi_chu)

        # Tùy chọn xử lý thuộc tính cũ
        self.chk_delete_old_fields = QCheckBox("Xóa toàn bộ thuộc tính cũ của layer trước khi thêm thuộc tính mới", self)
        self.chk_delete_old_fields.setChecked(False)
        scroll_layout.addWidget(self.chk_delete_old_fields)

        # 2. Khu vực danh sách Layer
        scroll_layout.addWidget(QLabel("Chọn các layer cần thêm trường dữ liệu:"))
        self.list_widget = QListWidget(self)
        self.layer_map = {}
        
        for layer in vector_layers:
            item = QListWidgetItem(layer.name())
            item.setFlags(item.flags() | FLAG_CHECKABLE)
            item.setCheckState(CHK_CHECKED)
            item.setData(USER_ROLE, layer.id()) 
            self.list_widget.addItem(item)
            self.layer_map[layer.id()] = layer
            
        scroll_layout.addWidget(self.list_widget)

        # 3. Các nút chức năng chọn nhanh
        btn_layout = QHBoxLayout()
        btn_select_all = QPushButton("Chọn tất cả")
        btn_deselect_all = QPushButton("Bỏ chọn tất cả")
        
        btn_select_all.clicked.connect(self.select_all)
        btn_deselect_all.clicked.connect(self.deselect_all)
        
        btn_layout.addWidget(btn_select_all)
        btn_layout.addWidget(btn_deselect_all)
        scroll_layout.addLayout(btn_layout)

        scroll.setWidget(scroll_content)
        layout.addWidget(scroll)

        # 4. Nút OK / Cancel 
        self.button_box = QDialogButtonBox(BTN_OK | BTN_CANCEL)
        self.button_box.accepted.connect(self.accept)
        self.button_box.rejected.connect(self.reject)
        layout.addWidget(self.button_box)

    def select_all(self):
        for i in range(self.list_widget.count()):
            self.list_widget.item(i).setCheckState(CHK_CHECKED)

    def deselect_all(self):
        for i in range(self.list_widget.count()):
            self.list_widget.item(i).setCheckState(CHK_UNCHECKED)

    def get_selected_layers(self):
        selected = []
        for i in range(self.list_widget.count()):
            item = self.list_widget.item(i)
            if item.checkState() == CHK_CHECKED:
                layer_id = item.data(USER_ROLE)
                selected.append(self.layer_map[layer_id])
        return selected

# Hàm hỗ trợ lấy Index cột (Khắc phục lỗi Shapefile cắt tên thành 10 ký tự)
def get_field_idx(layer, field_name):
    idx = layer.fields().indexOf(field_name)
    if idx == -1:
        # Thử tìm tên cột bị cắt bớt 10 ký tự (Đặc thù của file .shp)
        idx = layer.fields().indexOf(field_name[:10])
    return idx


# --- EXCEL: xuất/nhập thuộc tính (nhúng; cần openpyxl) ---
try:
    from qgis.core import NULL as QGIS_NULL
except ImportError:
    QGIS_NULL = object()

try:
    from openpyxl import Workbook, load_workbook
except ImportError:
    Workbook = None
    load_workbook = None

_FID_HEADER = "qgis_fid"


def _openpyxl_available():
    return Workbook is not None and load_workbook is not None


def sanitize_filename(value):
    text = str(value).strip()
    if not text:
        return "layer"
    for ch in '<>:"/\\|?*':
        text = text.replace(ch, "_")
    text = text.replace(" ", "_").replace("-", "_")
    return text.strip("._") or "layer"


def _excel_scalar(val):
    if val is None or val == QGIS_NULL:
        return None
    if isinstance(val, QVariant) and val.isNull():
        return None
    if isinstance(val, QDateTime):
        if val.isValid():
            return val.toPyDateTime()
        return None
    if isinstance(val, QDate):
        if val.isValid():
            return val.toPyDate()
        return None
    return val


def _write_layer_workbook(layer):
    wb = Workbook()
    ws = wb.active
    ws.title = "thuoc_tinh"
    fields = layer.fields()
    headers = [_FID_HEADER] + [fields.at(i).name() for i in range(fields.count())]
    ws.append(headers)
    for feat in layer.getFeatures():
        row = [feat.id()]
        for i in range(fields.count()):
            row.append(_excel_scalar(feat.attribute(i)))
        ws.append(row)
    return wb


def _save_workbook(wb, path):
    os.makedirs(os.path.dirname(path) or ".", exist_ok=True)
    wb.save(path)


def export_layers_attributes_excel(layers, output_path):
    if not _openpyxl_available():
        return "", "Thiếu thư viện openpyxl. Cài trong Python của QGIS: pip install openpyxl"
    if not layers:
        return "", "Không có lớp nào được chọn."
    n = len(layers)
    if n == 1:
        layer = layers[0]
        path = output_path.strip().rstrip("/\\")
        if not path:
            path = os.path.join(".", sanitize_filename(layer.name()) + ".xlsx")
        elif os.path.isdir(path):
            path = os.path.join(path, sanitize_filename(layer.name()) + ".xlsx")
        else:
            base_dir, fname = os.path.split(path)
            stem, ext = os.path.splitext(fname)
            if not stem:
                path = os.path.join(base_dir or ".", sanitize_filename(layer.name()) + ".xlsx")
        if not path.lower().endswith(".xlsx"):
            path = path + ".xlsx" if not path.endswith((".xlsx", ".zip")) else path
        if not path.lower().endswith(".xlsx"):
            path = os.path.splitext(path)[0] + ".xlsx"
        wb = _write_layer_workbook(layer)
        _save_workbook(wb, path)
        return path, None
    zip_path = output_path.strip().rstrip("/\\")
    if not zip_path:
        zip_path = os.path.join(".", f"{sanitize_filename(layers[0].name())}_{n}lop.zip")
    elif os.path.isdir(zip_path):
        zip_path = os.path.join(zip_path, f"{sanitize_filename(layers[0].name())}_{n}lop.zip")
    if not zip_path.lower().endswith(".zip"):
        zip_path = os.path.splitext(zip_path)[0] + ".zip"
    used_names = {}
    with tempfile.TemporaryDirectory(prefix="hsg_excel_") as tmp:
        files_to_zip = []
        for layer in layers:
            base = sanitize_filename(layer.name())
            k = used_names.get(base, 0)
            used_names[base] = k + 1
            fname = f"{base}.xlsx" if k == 0 else f"{base}_{k + 1}.xlsx"
            fpath = os.path.join(tmp, fname)
            wb = _write_layer_workbook(layer)
            _save_workbook(wb, fpath)
            files_to_zip.append((fpath, fname))
        with zipfile.ZipFile(zip_path, "w", zipfile.ZIP_DEFLATED) as zf:
            for fpath, arcname in files_to_zip:
                zf.write(fpath, arcname)
    return zip_path, None


def _header_map(ws):
    row = next(ws.iter_rows(min_row=1, max_row=1, values_only=True), None)
    if not row:
        return {}
    return {str(c).strip(): i for i, c in enumerate(row) if c is not None and str(c).strip()}


def _apply_sheet_to_layer(layer, ws):
    hmap = _header_map(ws)
    if _FID_HEADER not in hmap:
        return -1
    fid_col = hmap[_FID_HEADER]
    fields = layer.fields()
    field_names = [fields.at(i).name() for i in range(fields.count())]
    col_to_field_idx = {}
    for name in field_names:
        for hdr, cidx in hmap.items():
            if hdr == _FID_HEADER:
                continue
            if hdr == name or hdr == name[:10]:
                fi = fields.indexOf(name)
                if fi >= 0:
                    col_to_field_idx[cidx] = fi
                break
    layer.startEditing()
    pr = layer.dataProvider()
    updated = 0
    rows = list(ws.iter_rows(min_row=2, values_only=True))
    for row in rows:
        if row is None or fid_col >= len(row):
            continue
        raw_fid = row[fid_col]
        if raw_fid is None or str(raw_fid).strip() == "":
            continue
        try:
            fid = int(raw_fid)
        except (TypeError, ValueError):
            continue
        if not layer.getFeature(fid).isValid():
            continue
        attrs = {}
        for cidx, fidx in col_to_field_idx.items():
            if cidx >= len(row):
                continue
            val = row[cidx]
            attrs[fidx] = val
        if attrs:
            pr.changeAttributeValues({fid: attrs})
            updated += 1
    if not layer.commitChanges():
        layer.rollBack()
        return -2
    return updated


def import_layers_attributes_excel(layers, input_path, log_print=print):
    if not _openpyxl_available():
        return "Thiếu thư viện openpyxl. Cài trong Python của QGIS: pip install openpyxl"
    if not layers:
        return "Không có lớp nào được chọn."
    if not input_path or not os.path.isfile(input_path):
        return "File không tồn tại."
    ext = os.path.splitext(input_path)[1].lower()
    if ext == ".zip":
        groups = defaultdict(list)
        for lyr in layers:
            groups[sanitize_filename(lyr.name())].append(lyr)
        with tempfile.TemporaryDirectory(prefix="hsg_excel_in_") as tmp:
            with zipfile.ZipFile(input_path, "r") as zf:
                zf.extractall(tmp)
            stem_to_path = {}
            for root, _, files in os.walk(tmp):
                for fn in files:
                    if not fn.lower().endswith(".xlsx"):
                        continue
                    stem = os.path.splitext(fn)[0]
                    stem_to_path[stem] = os.path.join(root, fn)
            matched = 0
            for base, lyr_list in groups.items():
                for i, layer in enumerate(lyr_list):
                    stem_key = base if i == 0 else f"{base}_{i + 1}"
                    fpath = stem_to_path.get(stem_key)
                    if not fpath:
                        log_print(
                            f"  − Trong ZIP không có file tương ứng lớp «{layer.name()}» (mong đợi: {stem_key}.xlsx)."
                        )
                        continue
                    wb = load_workbook(fpath, read_only=True, data_only=True)
                    try:
                        ws = wb.active
                        n = _apply_sheet_to_layer(layer, ws)
                        if n == -2:
                            log_print(
                                f"  − Không lưu được thay đổi cho lớp {layer.name()} (từ {os.path.basename(fpath)})."
                            )
                        elif n < 0:
                            log_print(f"  − File {os.path.basename(fpath)} thiếu cột bắt buộc '{_FID_HEADER}'.")
                        else:
                            matched += 1
                            log_print(
                                f"  + Đã cập nhật {n} đối tượng từ {os.path.basename(fpath)} → {layer.name()}"
                            )
                    finally:
                        wb.close()
            if matched == 0:
                return "Không khớp được file .xlsx nào trong ZIP với các lớp đã chọn (tên file = tên lớp đã làm sạch)."
        return None
    if ext != ".xlsx":
        return "Chỉ hỗ trợ file .xlsx hoặc .zip."
    wb = load_workbook(input_path, read_only=True, data_only=True)
    try:
        ws = wb.active
        if len(layers) == 1:
            n = _apply_sheet_to_layer(layers[0], ws)
            if n == -2:
                return f"Không lưu được thay đổi cho lớp {layers[0].name()}."
            if n < 0:
                return f"File thiếu cột bắt buộc '{_FID_HEADER}'."
            log_print(f"  + Đã cập nhật {n} đối tượng cho lớp {layers[0].name()}.")
            return None
        stem = sanitize_filename(os.path.splitext(os.path.basename(input_path))[0])
        target = None
        for lyr in layers:
            if sanitize_filename(lyr.name()) == stem or lyr.name() == stem:
                target = lyr
                break
        if target is None:
            return (
                "Nhiều lớp được chọn: hãy dùng file ZIP từ bước xuất, hoặc chọn đúng một lớp, "
                "hoặc đặt tên file .xlsx trùng tên lớp (đã làm sạch)."
            )
        n = _apply_sheet_to_layer(target, ws)
        if n == -2:
            return f"Không lưu được thay đổi cho lớp {target.name()}."
        if n < 0:
            return f"File thiếu cột bắt buộc '{_FID_HEADER}'."
        log_print(f"  + Đã cập nhật {n} đối tượng cho lớp {target.name()}.")
        return None
    finally:
        wb.close()


def add_fields_and_data():
    all_layers = QgsProject.instance().mapLayers().values()
    vector_layers = [lyr for lyr in all_layers if lyr.type() == QgsMapLayerType.VectorLayer]
    
    if not vector_layers:
        print("Không tìm thấy layer Vector nào trong dự án.")
        return

    dialog = LayerSelectionDialog(vector_layers)
    
    if hasattr(dialog, 'exec'):
        is_accepted = dialog.exec()
    else:
        is_accepted = dialog.exec_()

    if is_accepted:
        # Lấy dữ liệu từ giao diện
        ma_tt_qh = dialog.input_ma_tt.text().strip()
        ma_hs_qh = dialog.input_ma_hs.text().strip()
        ma_dt_goc = dialog.input_ma_dt.text().strip()
        ten_dt = dialog.input_ten_dt.text().strip()
        phan_loai = dialog.input_phan_loai.text().strip()
        ghi_chu = dialog.input_ghi_chu.text().strip()
        delete_old_fields = dialog.chk_delete_old_fields.isChecked()
        
        selected_layers = dialog.get_selected_layers()
        if not selected_layers:
            print("Bạn chưa chọn layer nào để xử lý.")
            return
            
        print(f"Bắt đầu xử lý {len(selected_layers)} layer đã chọn...\n")

        # Cấu trúc trường theo chuẩn (Độ dài được khai báo đúng quy định)
        fields_to_add = [
            QgsField("maThongTinQH", QVariant.String, len=15),
            QgsField("maHoSoQH", QVariant.String, len=15),
            QgsField("maDoiTuong", QVariant.String, len=100),
            QgsField("tenDoiTuong", QVariant.String, len=100),
            QgsField("phanLoai", QVariant.String, len=250),
            QgsField("ghiChu", QVariant.String, len=250)
        ]

        for layer in selected_layers:
            print(f"Đang cập nhật: {layer.name()}")
            layer.startEditing()
            pr = layer.dataProvider()

            # 1. Tùy chọn xóa toàn bộ trường cũ
            if delete_old_fields:
                old_field_count = len(layer.fields())
                if old_field_count > 0:
                    pr.deleteAttributes(list(range(old_field_count)))
                    layer.updateFields()
                    print(f"  + Đã xóa {old_field_count} thuộc tính cũ")

            # 2. Thêm trường nếu chưa có
            existing_fields = layer.fields().names()
            new_fields = [f for f in fields_to_add if f.name() not in existing_fields and f.name()[:10] not in existing_fields]
            if new_fields:
                pr.addAttributes(new_fields)
                layer.updateFields()

            # 3. Lấy Index chính xác của cột (Kể cả khi bị cắt 10 ký tự)
            idx_maThongTinQH = get_field_idx(layer, "maThongTinQH")
            idx_maHoSoQH = get_field_idx(layer, "maHoSoQH")
            idx_maDoiTuong = get_field_idx(layer, "maDoiTuong")
            idx_tenDoiTuong = get_field_idx(layer, "tenDoiTuong")
            idx_phanLoai = get_field_idx(layer, "phanLoai")
            idx_ghiChu = get_field_idx(layer, "ghiChu")

            # 4. Cập nhật dữ liệu
            update_dict = {}
            for feat in layer.getFeatures():
                obj_id = feat.id()
                ma_doi_tuong = ma_dt_goc
                
                # Chỉ cập nhật các cột tìm thấy
                attr_map = {}
                if idx_maThongTinQH != -1: attr_map[idx_maThongTinQH] = ma_tt_qh
                if idx_maHoSoQH != -1: attr_map[idx_maHoSoQH] = ma_hs_qh
                if idx_maDoiTuong != -1: attr_map[idx_maDoiTuong] = ma_doi_tuong
                if idx_tenDoiTuong != -1 and ten_dt: attr_map[idx_tenDoiTuong] = ten_dt
                if idx_phanLoai != -1 and phan_loai: attr_map[idx_phanLoai] = phan_loai
                if idx_ghiChu != -1 and ghi_chu: attr_map[idx_ghiChu] = ghi_chu
                
                update_dict[obj_id] = attr_map

            pr.changeAttributeValues(update_dict)
            layer.commitChanges()
            print(f"  + Hoàn tất: {layer.name()}")

        print("\n--- ĐÃ CHẠY XONG TOÀN BỘ YÊU CẦU ---")
        QMessageBox.information(
            None,
            "Hoàn tất",
            "Quá trình đã hoàn tất.\nPhần mềm được phát triển bởi LEDAT.\nCảm ơn đã sử dụng."
        )
    else:
        print("Đã hủy thao tác.")


if __name__ == "__main__":
    add_fields_and_data()